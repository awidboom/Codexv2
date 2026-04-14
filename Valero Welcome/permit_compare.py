from __future__ import annotations

import re
from collections import Counter, OrderedDict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PyPDF2 import PdfReader


ROOT = Path(__file__).resolve().parent
OLD_PDF = ROOT / "09100062-103 Permit.pdf"
NEW_PDF = ROOT / "Preliminary Draft Permit.pdf"
XLSX_OUT = ROOT / "permit_comparison_analysis.xlsx"
XLSX_FALLBACK_OUT = ROOT / "permit_comparison_analysis_updated.xlsx"
MD_OUT = ROOT / "permit_comparison_summary.md"
APPLICATION_DIR = ROOT / "Permit applications"
WORKFLOW_XLSX = ROOT / "Valero_Welcome_Permit_Comparison.xlsx"

MAJOR_APP = "Major Amendment DDGS BH Maintenance_2025.11.10.pdf"
MINOR_7277_APP = "Valero Welcome Minor Amendment 7277_2025.02.27.pdf"
MINOR_DDGS_APP = "Valero Welcome Air Permit Minor Mod - DDGS Drag Replacement.pdf"
MINOR_TO_APP = "Valero Welcome Air Permit Minor Mod - TO Inspection.pdf"
ALL_APPS = "; ".join([MAJOR_APP, MINOR_7277_APP, MINOR_DDGS_APP, MINOR_TO_APP])

HEADERS = [
    "Comparison ID",
    "Old Permit Section",
    "New Permit Section",
    "Subject Tag Primary",
    "Subject Tag Secondary",
    "Old Text",
    "New Text",
    "Change Type",
    "Initial Analysis",
    "Regulatory or Compliance Impact",
    "Operational Impact",
    "Risk Level",
    "Recommended Follow-Up",
    "Consultant Conclusion",
    "Requested Change Status",
    "Application Source",
    "Application Basis",
    "Review Priority",
    "Validation Basis",
    "Evidence Confidence",
]

APPLICATION_FILES = [
    (
        MAJOR_APP,
        "Current draft application processed with the 2025-2026 permit package",
        "Request to modify STRU 11/TREA 2 language so EQUI 243, 18, and 191 can route DDGS to EQUI 218 during TREA 2 maintenance for up to 100 hours per year.",
        "This is the only application in the set that directly requests the new STRU 11 maintenance-routing scenario. It does not describe the ambient-boundary program, modeling-framework deletions, STRU 13/TREA 4 threshold changes, or broader testing rewrites.",
    ),
    (
        MINOR_7277_APP,
        "Concurrent minor amendment processed with the same overall permit package",
        "Replace DDGS Inclined Drag (EQUI 222) with a higher-throughput drag while keeping STRU 11 limits and throughput unchanged.",
        "Important cross-check for the current draft because it expressly says STRU 11 emission limits and throughput were not being changed. That makes the draft's new 178.8 tph STRU 11 limit look agency-driven rather than applicant-requested.",
    ),
    (
        MINOR_DDGS_APP,
        "Earlier minor amendment rolled into the current permit history",
        "Replace the DDGS top fill drag conveyor and keep STRU 11 emission limits and throughput unchanged.",
        "Useful historical context because it also states that STRU 11 emission limits and throughput were not being modified. It reinforces that the new draft throughput tightening was not a longstanding applicant request.",
    ),
    (
        MINOR_TO_APP,
        "Earlier minor amendment processed in the same broader permit package",
        "Align TREA 16 internal inspection frequency with TREA 11 and update the supporting vendor-letter reference.",
        "This application is narrow and administrative. It is useful mainly to confirm that unrelated draft changes to modeling, STRU 13/TREA 4, and STRU 14 were not requested through the TO inspection package.",
    ),
]

REQUEST_META = {
    "PC-001": ("Related to requested change", MAJOR_APP, "The new permit number and major-amendment posture are consistent with the current major amendment application and the bundled permit package, but do not represent a separate review issue.", "Routine"),
    "PC-002": ("Related to requested change", ALL_APPS, "The draft applications table reflects that MPCA appears to be processing the three minor amendments and the major amendment together, plus an agency reopening entry. That expanded processing posture matters for scope, but it does not by itself make every draft change applicant-requested.", "Watch"),
    "PC-003": ("Requested", MAJOR_APP, "The major amendment narrative expressly reframes the project around routing DDGS to EQUI 218 during TREA 2 maintenance, so the revised facility description is expected.", "Routine"),
    "PC-004": ("Not identified in application", ALL_APPS, "None of the four applications request retiring the issued permit's PM10 equivalent-or-better-dispersion trigger language.", "Highlight"),
    "PC-005": ("Not identified in application", ALL_APPS, "None of the four applications request deleting the issued permit's detailed PM10 modeling and protocol framework.", "Highlight"),
    "PC-006": ("Not identified in application", ALL_APPS, "None of the four applications request a General Public Preclusion Plan, effective-fence-line mapping, or new ambient-boundary controls.", "Highlight"),
    "PC-007": ("Not identified in application", ALL_APPS, "None of the four applications identify a new ambient-boundary breach documentation and reporting program.", "Highlight"),
    "PC-008": ("Not identified in application", ALL_APPS, "None of the four applications request new fencing, camera, or controlled-access obligations.", "Highlight"),
    "PC-009": ("Not identified in application", ALL_APPS, "Daily patrols and remote-monitoring duties are not described in any of the four application packages reviewed.", "Highlight"),
    "PC-010": ("Not identified in application", ALL_APPS, "None of the four applications request the new spill-cleanup and public-nuisance language.", "Highlight"),
    "PC-011": ("Related to requested change", MAJOR_APP, "AOS 1 labeling and the companion TREA 2 control-language clarification are implementation details tied to the requested maintenance-routing change.", "Routine"),
    "PC-012": ("Requested", MAJOR_APP, "The major amendment expressly requests allowing EQUI 243, 18, and 191 to route DDGS to EQUI 218 while TREA 2 is down for maintenance.", "Routine"),
    "PC-013": ("Not identified in application", ALL_APPS, "None of the four applications explain or request a permit structure in which the carried-forward PM and opacity limits are expressly limited to AOS 1 without parallel AOS 2 limits.", "Highlight"),
    "PC-014": ("Related to requested change", MAJOR_APP, "The major amendment expressly references a 100-hour annual cap; the added tracking records are implementation details for that requested limit.", "Watch"),
    "PC-015": ("Not identified in application", f"{MAJOR_APP}; {MINOR_7277_APP}; {MINOR_DDGS_APP}", "The major amendment requests the 100-hour maintenance-routing scenario but does not request lowering the STRU 11 throughput limit. Both DDGS drag minor amendments also expressly state that STRU 11 emission limits and throughput were not being changed, so the new 178.8 tph limit does not trace back to any of those applications.", "Highlight"),
    "PC-016": ("Legacy or structural only", MAJOR_APP, "This block is largely renumbering and relocation needed to accommodate the new STRU 11 provisions.", "Routine"),
    "PC-017": ("Related to requested change", MAJOR_APP, "Labeling the periodic STRU 11 tests as AOS 1 is a scenario clarification tied to the requested AOS structure.", "Routine"),
    "PC-018": ("Not identified in application", ALL_APPS, "None of the four applications request changing the particulate test methods for STRU 10-12.", "Watch"),
    "PC-019": ("Not identified in application", ALL_APPS, "None of the four applications identify the reduced STRU 13 beer-feed limit as a requested change.", "Watch"),
    "PC-020": ("Not identified in application", ALL_APPS, "None of the four applications describe resetting STRU 13 periodic testing due dates to October 30, 2028.", "Watch"),
    "PC-022": ("Not identified in application", ALL_APPS, "None of the four applications request the revised AOS 1 fresh-water CAM threshold for the CO2 scrubber.", "Watch"),
    "PC-023": ("Not identified in application", ALL_APPS, "None of the four applications request the tighter 145 gpm recirculating-scrubber-water CAM threshold.", "Highlight"),
    "PC-024": ("Legacy or structural only", "None", "This appears to be an ordering change only and is not tied to a distinct application request.", "Routine"),
    "PC-025": ("Legacy or structural only", "None", "This appears to be CAM reordering rather than a requested substantive revision.", "Routine"),
    "PC-026": ("Not identified in application", ALL_APPS, "None of the four applications request the simplified STRU 14 periodic SO2 testing language.", "Highlight"),
    "PC-027": ("Legacy or structural only", "None", "The parenthetical expiration date is an agency clarification, not a separately requested permit change.", "Routine"),
    "PC-028": ("Not identified in application", ALL_APPS, "None of the four applications request deleting the issued permit's NOx and PM2.5 equivalent-or-better-dispersion trigger and protocol framework.", "Highlight"),
}

WORKFLOW_USE_NOTE = (
    "The uploaded workflow workbook is useful as a screening tool, but manual spot-checking confirms "
    "that its downstream summary tabs can overcall removals or additions when requirements are retained, "
    "renumbered, regrouped, or affected by OCR artifacts. Final client advice should rely on the "
    "Detailed Comparison, Matched Conditions, and QA Spot Checks tabs rather than on the workflow "
    "Removed & Added or Equipment Summaries tabs alone."
)

MANUAL_QA_FINDINGS = [
    {
        "QA ID": "QA-001",
        "Workflow Sheet": "Removed & Added",
        "Subject Item": "COMG 2",
        "Requirement / Topic": "5.2.9",
        "Workflow Claim": "Listed as removed from the draft permit.",
        "Manual Permit Check Result": "False positive. Draft requirement 5.2.9 is still present and materially the same as the issued permit.",
        "Old Permit Evidence": "Issued permit 5.2.9 contains the STRU 13/TREA 4 AOS 1 and AOS 2 HAP emission-factor language.",
        "Draft Permit Evidence": "Draft permit 5.2.9 retains the same COMG 2 source-specific HAP emission-factor condition at the same requirement number.",
        "QA Conclusion": "Workflow false positive",
        "Recommended Use": "Do not treat as a deletion. Review STRU 13/TREA 4 on its own merits rather than relying on the Removed & Added tab.",
    },
    {
        "QA ID": "QA-002",
        "Workflow Sheet": "Removed & Added",
        "Subject Item": "COMG 2",
        "Requirement / Topic": "5.2.10",
        "Workflow Claim": "Listed as removed or collapsed as a duplicate in the draft permit.",
        "Manual Permit Check Result": "False positive. Draft requirement 5.2.10 remains in place, and draft 5.2.16 also still contains the separate concurrent-testing reset protocol.",
        "Old Permit Evidence": "Issued permit 5.2.10 contains the concurrent VOC/HAP reset protocol language.",
        "Draft Permit Evidence": "Draft permit 5.2.10 remains present at the same requirement number; differences are limited to formatting/OCR cleanup.",
        "QA Conclusion": "Workflow false positive",
        "Recommended Use": "Do not characterize this as a removed condition without a direct permit citation.",
    },
    {
        "QA ID": "QA-003",
        "Workflow Sheet": "Removed & Added",
        "Subject Item": "COMG 2",
        "Requirement / Topic": "5.2.11",
        "Workflow Claim": "Listed as removed from the draft permit.",
        "Manual Permit Check Result": "False positive. Draft requirement 5.2.11 is still present and materially unchanged.",
        "Old Permit Evidence": "Issued permit 5.2.11 contains the STRU 14 / TREA 11 / TREA 16 HAP emission-factor and combustion-HAP methodology language.",
        "Draft Permit Evidence": "Draft permit 5.2.11 retains the same requirement at the same number; the observed differences are OCR spacing only.",
        "QA Conclusion": "Workflow false positive",
        "Recommended Use": "Treat the workbook flag as a screening error, not as a substantive permit change.",
    },
    {
        "QA ID": "QA-004",
        "Workflow Sheet": "Equipment Summaries",
        "Subject Item": "COMG 2",
        "Requirement / Topic": "Summary narrative",
        "Workflow Claim": "Summary states that COMG-level source-specific HAP emission-factor provisions were removed from COMG 2.",
        "Manual Permit Check Result": "Overstated. Manual review of COMG 2 requirements 5.2.1 through 5.2.29 did not confirm the claimed removals in the summary narrative.",
        "Old Permit Evidence": "Issued permit COMG 2 includes 29 requirements, including 5.2.9, 5.2.10, and 5.2.11.",
        "Draft Permit Evidence": "Draft permit COMG 2 also includes 29 requirements, and the specific COMG 2 conditions cited as removed remain in the draft.",
        "QA Conclusion": "Workflow summary overcall",
        "Recommended Use": "Use Equipment Summaries only as a screening aid. Confirm conclusions against the permit text before carrying them into client advice.",
    },
]

REQ_PATTERN = re.compile(r"^(\d+\.\d+\.\d+)\s+")
SPLIT_REQ_PATTERN = re.compile(r"^((?:\d+\.){2}\d)\s+(\d)\s+(.*)$")
IGNORE_PREFIXES = (
    "Permit Issued:",
    "Permit issued:",
    "Permit Expires:",
    "Permit expires:",
    "Requirement number",
    "Requirement n umber",
    "Requirement and citation",
    "Requirem ent and c itation",
    "Page ",
)
WORD_RE = re.compile(r"[A-Za-z][A-Za-z0-9.]*")
ALPHA_SPLIT_RE = re.compile(r"^[A-Za-z]+$")
CUSTOM_WORDS = {
    "particulate",
    "permittee",
    "micron",
    "operating",
    "operation",
    "emissions",
    "recirculating",
    "throughput",
    "performance",
    "compliance",
    "requirements",
    "monitoring",
    "recordkeeping",
    "temperature",
    "baghouse",
    "deviation",
    "commissioner",
    "pollutants",
    "conditions",
    "authorized",
    "ambient",
    "scrubber",
    "methanol",
    "acetaldehyde",
    "formaldehyde",
    "acrolein",
    "hexane",
    "worstcase",
    "shortterm",
    "throughputlimit",
    "throughputlimits",
}

OCR_FIXES = (
    (re.compile(r"\bth\s+is\b", re.IGNORECASE), "this"),
    (re.compile(r"\bth\s+e\b", re.IGNORECASE), "the"),
    (re.compile(r"\bfr\s+om\b", re.IGNORECASE), "from"),
    (re.compile(r"\bme\s+et\b", re.IGNORECASE), "meet"),
    (re.compile(r"\bte\s+st\b", re.IGNORECASE), "test"),
    (re.compile(r"\bfo\s+r\b", re.IGNORECASE), "for"),
    (re.compile(r"\bth\s+an\b", re.IGNORECASE), "than"),
    (re.compile(r"\bdu\s+e\b", re.IGNORECASE), "due"),
    (re.compile(r"\bmo\s+st\b", re.IGNORECASE), "most"),
    (re.compile(r"\bea\s+ch\b", re.IGNORECASE), "each"),
    (re.compile(r"\bworst\s*-\s*case\b", re.IGNORECASE), "worst-case"),
    (re.compile(r"\bworstcase\b", re.IGNORECASE), "worst-case"),
    (re.compile(r"\bshort\s*-\s*term\b", re.IGNORECASE), "short-term"),
    (re.compile(r"\bshortterm\b", re.IGNORECASE), "short-term"),
    (re.compile(r"\bhard\s*-\s*copy\b", re.IGNORECASE), "hard-copy"),
    (re.compile(r"\bthroughputlimits\b", re.IGNORECASE), "throughput limits"),
    (re.compile(r"\bthroughputlimit\b", re.IGNORECASE), "throughput limit"),
    (re.compile(r"\bthoughputlimits\b", re.IGNORECASE), "throughput limits"),
    (re.compile(r"\bthoughputlimit\b", re.IGNORECASE), "throughput limit"),
    (re.compile(r"\blimit\s+s\b", re.IGNORECASE), "limits"),
    (re.compile(r"\bdescribedat\b", re.IGNORECASE), "described at"),
    (re.compile(r"\btodocument\b", re.IGNORECASE), "to document"),
    (re.compile(r"\bmusttest\b", re.IGNORECASE), "must test"),
    (re.compile(r"\bbyMPCA\b"), "by MPCA"),
    (re.compile(r"\bfollow up\b", re.IGNORECASE), "follow-up"),
    (re.compile(r"\b([1-9])\s+(0\d{2}\.\d{4})\b"), r"\1\2"),
)


@dataclass
class RowSpec:
    comparison_id: str
    old_section: str
    new_section: str
    subject_primary: str
    subject_secondary: str
    old_text_getter: Callable[[], str]
    new_text_getter: Callable[[], str]
    change_type: str
    initial_analysis: str
    regulatory_impact: str
    operational_impact: str
    risk_level: str
    follow_up: str
    conclusion: str
    importance_rank: int


def normalize_line(raw: str) -> str:
    line = re.sub(r"\s+", " ", raw.replace("\r", " ")).strip()
    split_match = SPLIT_REQ_PATTERN.match(line)
    if split_match:
        line = f"{split_match.group(1)}{split_match.group(2)} {split_match.group(3)}"
    return line


def build_vocab(pdf_paths: list[Path]) -> set[str]:
    vocab = set(CUSTOM_WORDS)
    for pdf_path in pdf_paths:
        reader = PdfReader(str(pdf_path))
        for page in reader.pages:
            text = page.extract_text() or ""
            for word in WORD_RE.findall(text):
                vocab.add(word.lower())
    return vocab


def cleanup_text(text: str, vocab: set[str]) -> str:
    text = text.replace("–", "-").replace("—", "-").replace("‑", "-")
    text = re.sub(r"(?<=\d)\s*-\s*(?=[A-Za-z])", "-", text)
    text = re.sub(r"(?<=\d)\s*-\s*(?=\d)", "-", text)
    text = re.sub(r"(?<=\d)-\s+(?=\d)", "-", text)
    text = re.sub(r"(?<=\d)\.\s+(?=\d)", ".", text)
    text = re.sub(r"(?<=[A-Za-z0-9])\s+\((?=[A-Za-z0-9])", "(", text)
    text = re.sub(r"(?<=\d)\s+\)", ")", text)
    text = re.sub(r"\(\s+(?=[A-Za-z0-9])", "(", text)
    text = re.sub(r"\s+", " ", text).strip()

    tokens = re.findall(r"[A-Za-z]+|[^A-Za-z]+", text)
    healed: list[str] = []
    idx = 0
    while idx < len(tokens):
        current = tokens[idx]
        if (
            idx + 2 < len(tokens)
            and ALPHA_SPLIT_RE.match(current or "")
            and tokens[idx + 1] == " "
            and ALPHA_SPLIT_RE.match(tokens[idx + 2] or "")
        ):
            joined = f"{current}{tokens[idx + 2]}"
            if joined.lower() in vocab and len(joined) >= 3:
                healed.append(joined)
                idx += 3
                continue
        healed.append(current)
        idx += 1

    text = "".join(healed)
    for pattern, replacement in OCR_FIXES:
        text = pattern.sub(replacement, text)
    text = re.sub(r"\b([A-Za-z]{2,})\s*-\s*([A-Za-z]{2,})\b", r"\1-\2", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_requirements(pdf_path: Path) -> OrderedDict[str, str]:
    reader = PdfReader(str(pdf_path))
    reqs: OrderedDict[str, str] = OrderedDict()
    current_req: str | None = None
    buffer: list[str] = []

    for page in reader.pages:
        text = page.extract_text() or ""
        for raw in text.split("\n"):
            line = normalize_line(raw)
            if not line or any(line.startswith(prefix) for prefix in IGNORE_PREFIXES):
                continue
            match = REQ_PATTERN.match(line)
            if match:
                if current_req:
                    reqs[current_req] = cleanup_text(" ".join(buffer).strip(), VOCAB)
                current_req = match.group(1)
                buffer = [line[match.end() :].strip()]
            elif current_req:
                buffer.append(line)

    if current_req:
        reqs[current_req] = cleanup_text(" ".join(buffer).strip(), VOCAB)
    return reqs


VOCAB = build_vocab([OLD_PDF, NEW_PDF])
OLD_REQS = extract_requirements(OLD_PDF)
NEW_REQS = extract_requirements(NEW_PDF)


def req_text(reqs: OrderedDict[str, str], req_id: str) -> str:
    text = reqs.get(req_id, "")
    return f"{req_id} {text}".strip()


def req_group(reqs: OrderedDict[str, str], req_ids: list[str]) -> str:
    return "\n\n".join(req_text(reqs, req_id) for req_id in req_ids if req_id in reqs).strip()


def request_meta(comparison_id: str) -> tuple[str, str, str, str]:
    return REQUEST_META.get(
        comparison_id,
        ("Not identified in application", "None", "No application basis was mapped for this comparison item.", "Watch"),
    )


def expand_section_reference(section: str, reqs: OrderedDict[str, str]) -> list[str]:
    if not section or "No direct equivalent" in section:
        return []
    found = re.findall(r"\d+\.\d+\.\d+", section)
    if not found:
        return []

    ordered_keys = list(reqs.keys())
    unique_found: list[str] = []
    for req in found:
        if req in reqs and req not in unique_found:
            unique_found.append(req)

    if len(found) == 2 and "-" in section and all(req in reqs for req in found):
        start, end = found
        try:
            start_idx = ordered_keys.index(start)
            end_idx = ordered_keys.index(end)
        except ValueError:
            return unique_found
        if start_idx <= end_idx:
            return ordered_keys[start_idx : end_idx + 1]
        return ordered_keys[end_idx : start_idx + 1]

    return unique_found


def summarize_validation(row: RowSpec) -> tuple[str, str]:
    old_refs = expand_section_reference(row.old_section, OLD_REQS)
    new_refs = expand_section_reference(row.new_section, NEW_REQS)

    if old_refs and new_refs:
        if row.change_type == "Possible relocation or renumbering":
            return ("Direct grouped condition comparison with renumbering/relocation review", "High")
        return ("Direct old/new permit condition comparison", "High")
    if old_refs and row.change_type == "Deletion":
        return ("Direct issued-permit condition deletion check", "High")
    if new_refs and row.change_type == "Addition":
        return ("Direct draft-permit condition addition check", "High")
    if old_refs or new_refs:
        return ("Direct condition review with structural or narrative comparison", "Medium")
    return ("Narrative/manual permit-structure comparison", "Medium")


def normalize_match_text(text: str) -> str:
    text = cleanup_text(text, VOCAB).lower()
    text = re.sub(r"\b\d+\.\d+\.\d+\b", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def condition_similarity(old_text: str, new_text: str) -> float:
    return SequenceMatcher(None, normalize_match_text(old_text), normalize_match_text(new_text)).ratio()


def build_matched_conditions() -> list[dict[str, str | float]]:
    old_items = list(OLD_REQS.items())
    new_items = list(NEW_REQS.items())
    matches: list[dict[str, str | float]] = []
    i = 0
    j = 0

    def append_old_only(old_id: str, old_text: str) -> None:
        matches.append(
            {
                "Match Status": "Old only",
                "Old Permit Section": old_id,
                "New Permit Section": "",
                "Similarity": 0.0,
                "Old Text": f"{old_id} {old_text}".strip(),
                "New Text": "",
            }
        )

    def append_new_only(new_id: str, new_text: str) -> None:
        matches.append(
            {
                "Match Status": "New only",
                "Old Permit Section": "",
                "New Permit Section": new_id,
                "Similarity": 0.0,
                "Old Text": "",
                "New Text": f"{new_id} {new_text}".strip(),
            }
        )

    while i < len(old_items) or j < len(new_items):
        if i >= len(old_items):
            new_id, new_text = new_items[j]
            append_new_only(new_id, new_text)
            j += 1
            continue
        if j >= len(new_items):
            old_id, old_text = old_items[i]
            append_old_only(old_id, old_text)
            i += 1
            continue

        old_id, old_text = old_items[i]
        new_id, new_text = new_items[j]
        direct_score = condition_similarity(old_text, new_text)

        if old_id == new_id or direct_score >= 0.73:
            matches.append(
                {
                    "Match Status": "Matched",
                    "Old Permit Section": old_id,
                    "New Permit Section": new_id,
                    "Similarity": round(direct_score, 3),
                    "Old Text": f"{old_id} {old_text}".strip(),
                    "New Text": f"{new_id} {new_text}".strip(),
                }
            )
            i += 1
            j += 1
            continue

        best: tuple[float, int, int] | None = None
        for oi in range(i, min(i + 4, len(old_items))):
            for nj in range(j, min(j + 4, len(new_items))):
                score = condition_similarity(old_items[oi][1], new_items[nj][1]) - (0.03 * ((oi - i) + (nj - j)))
                if best is None or score > best[0]:
                    best = (score, oi, nj)

        if best and best[0] >= 0.77:
            _, oi, nj = best
            while i < oi:
                old_skip_id, old_skip_text = old_items[i]
                append_old_only(old_skip_id, old_skip_text)
                i += 1
            while j < nj:
                new_skip_id, new_skip_text = new_items[j]
                append_new_only(new_skip_id, new_skip_text)
                j += 1
            continue

        if old_id < new_id:
            append_old_only(old_id, old_text)
            i += 1
        else:
            append_new_only(new_id, new_text)
            j += 1

    for index, row in enumerate(matches, start=1):
        row["Match ID"] = f"MC-{index:04d}"
    return matches


def manual(text: str) -> Callable[[], str]:
    return lambda: text


def old_req_group(req_ids: list[str]) -> Callable[[], str]:
    return lambda: req_group(OLD_REQS, req_ids)


def new_req_group(req_ids: list[str]) -> Callable[[], str]:
    return lambda: req_group(NEW_REQS, req_ids)


def autofit(ws, widths: dict[int, int]) -> None:
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_rows() -> list[RowSpec]:
    old_cover = (
        "Air Individual Permit. Minor Amendment. Permit No. 09100062-103. "
        "Minor Amendment: December 30, 2025. This permit amendment supersedes Air Emission Permit No. 09100062-102."
    )
    new_cover = (
        "Preliminary Draft Air Individual Permit. Major Amendment. Permit No. 09100062-104. "
        "Major Amendment: [Action Issue Date]. This permit amendment supersedes Air Emission Permit No. 09100062-103."
    )
    old_apps = (
        "Permit applications table lists Minor Amendment received 03/05/2025 under Action 09100062-103, "
        "plus earlier 2023 and 2021 actions under Permit 09100062-102 and reissuance/reopening under 09100062-101."
    )
    new_apps = (
        "Permit applications table adds Major Amendment received 12/04/2025 and MPCA Reopening dated 01/27/2026 under Action 09100062-104, "
        "while carrying forward the 03/05/2025 minor amendment and prior actions."
    )
    old_fac = (
        "Facility description states the minor amendment is for removing a drag conveyor and installing a new drag conveyor in its place at an increased throughput."
    )
    new_fac = (
        "Facility description states the major amendment adds an alternate operating scenario at STRU 11 to allow 100 hours of DDGS binning operations from EQUIs 18, 191, and 243 even when TREA 2 is undergoing maintenance."
    )

    rows = [
        RowSpec(
            "PC-001",
            "Cover Page / Permit Overview",
            "Cover Page / Permit Overview",
            "Administrative",
            "General Conditions",
            manual(old_cover),
            manual(new_cover),
            "Revision",
            "The draft changes the permit from an issued minor amendment under Permit 09100062-103 to a preliminary major amendment under Permit 09100062-104. That is more than a formatting change because it resets the procedural context for review and confirms the new draft would supersede the currently issued permit if finalized. The placeholders on the draft cover also indicate the agency has not yet finalized issuance details.",
            "This does not by itself change day-to-day compliance obligations, but it confirms the document under review is a proposed replacement permit rather than a simple redline to the issued permit.",
            "No direct operational effect. The practical effect is on review posture, permit comments, and how the client should frame any agency response.",
            "Low",
            "No action needed, but cite the correct permit number and action type in any client comments.",
            "Minor changes, likely acceptable",
            20,
        ),
        RowSpec(
            "PC-002",
            "1. Permit applications table",
            "1. Permit applications table",
            "Administrative",
            "Compliance Schedule",
            manual(old_apps),
            manual(new_apps),
            "Revision",
            "The applications table adds the December 4, 2025 major amendment application and a January 27, 2026 MPCA reopening under the new action number. It also updates some legacy receipt dates to a cleaner format. The substantive point is that the draft is tied to both a requested amendment and an agency reopening, which suggests the agency may be revisiting more than just the requested operational change.",
            "The added reopening entry can expand the scope of issues the agency may touch in the final permit. That matters for comment strategy because not every change necessarily flows only from the applicant's requested amendment.",
            "No immediate operational effect, but the combined amendment and reopening posture can enlarge the range of permit revisions that may be imposed.",
            "Low",
            "Confirm with agency or permit writer whether all major revisions in the draft are tied to the application, the reopening, or both.",
            "Minor changes, likely acceptable",
            19,
        ),
        RowSpec(
            "PC-003",
            "3. Facility description",
            "3. Facility description",
            "Facility Description",
            "Operational Limits",
            manual(old_fac),
            manual(new_fac),
            "Revision",
            "The facility description shifts from a drag-conveyor replacement narrative to a new major amendment centered on an alternate operating scenario at STRU 11. That is a real scope change, not an editorial cleanup, because the draft now frames the amendment around temporary direct venting during baghouse maintenance. The rest of the permit needs to be read in light of that revised project description.",
            "The project description becomes the context for the new operating scenario, new hour cap, and the revised STRU 11 limits and recordkeeping. If the description is inaccurate or incomplete, that can create interpretation issues later.",
            "This change can affect how maintenance outages are managed and how DDGS handling is scheduled while TREA 2 is unavailable.",
            "Medium",
            "Check consistency with current operations and verify the description matches the application narrative and intended maintenance practice.",
            "Significant changes",
            6,
        ),
        RowSpec(
            "PC-004",
            "5.1.57-5.1.58",
            "No direct equivalent",
            "Modeling",
            "Ambient Conditions",
            old_req_group(["5.1.57", "5.1.58"]),
            manual("No direct equivalent requirement appears in the draft. The draft shifts to a General Public Preclusion Plan framework instead of retaining these PM10 EBD trigger provisions."),
            "Deletion",
            "The issued permit contains explicit PM10 equivalent-or-better-dispersion trigger language that distinguishes changes not requiring modeling from changes that do. Those trigger provisions are removed from the draft as stand-alone requirements. That is meaningful because the old permit gave the source and reviewer a clearer front-end screen for when PM10 dispersion work had to be refreshed.",
            "Removing the explicit EBD trigger language can change how future modifications are evaluated for PM10 dispersion review. Even if the agency intends Appendix B and the new ambient-boundary controls to cover the same ground, the compliance roadmap is not the same.",
            "The change does not immediately alter current operations, but it could change future permitting strategy for modifications that affect modeled PM10 parameters.",
            "Medium",
            "Confirm with agency whether the EBD trigger concept is intentionally being retired or relocated and whether future PM10 changes will be evaluated under a different modeling approach.",
            "Significant changes",
            10,
        ),
        RowSpec(
            "PC-005",
            "5.1.59-5.1.65",
            "No direct equivalent",
            "Modeling",
            "Modeling",
            old_req_group(["5.1.59", "5.1.60", "5.1.61", "5.1.62", "5.1.63", "5.1.64", "5.1.65"]),
            manual("No one-for-one replacement. The draft does not retain the issued permit's stand-alone EBD submittal, content, baseline update, refined modeling trigger, or protocol deadline requirements."),
            "Deletion",
            "The issued permit includes a full PM10 EBD and refined-modeling framework: submittal content, outdated-baseline treatment, equivalency demonstration, refined-modeling triggers, and protocol deadlines. Those detailed modeling mechanics are not preserved as direct requirements in the draft. That is a substantive structural deletion because it removes explicit permit text that previously governed future PM10 modeling work.",
            "This could reduce clarity on what must be submitted before making changes that affect modeled PM10 parameters. It also shifts leverage from a permit-defined procedure toward case-by-case interpretation unless equivalent obligations appear elsewhere outside the permit text.",
            "No immediate operating change, but future projects could face more uncertainty over modeling expectations and timing.",
            "High",
            "Flag for client discussion and confirm whether the agency intends another enforceable document to replace the deleted PM10 modeling protocol language.",
            "Significant changes",
            5,
        ),
        RowSpec(
            "PC-028",
            "5.1.67-5.1.77",
            "No direct equivalent",
            "Modeling",
            "Ambient Conditions",
            old_req_group(["5.1.67", "5.1.68", "5.1.69", "5.1.70", "5.1.71", "5.1.72", "5.1.73", "5.1.74", "5.1.75", "5.1.76", "5.1.77"]),
            manual("No one-for-one replacement. The draft does not retain the issued permit's stand-alone NOx/PM2.5 EBD submittal, content, baseline update, refined modeling trigger, or protocol deadline requirements."),
            "Deletion",
            "Beyond the PM10 provisions, the issued permit also contains a separate NOx and PM2.5 equivalent-or-better-dispersion framework covering triggers, reissuance updates, required submittal content, outdated-baseline analysis, refined-modeling triggers, and protocol/report deadlines. Those NOx/PM2.5 requirements do not appear as stand-alone permit conditions in the draft. This is a meaningful deletion because it removes an additional pollutant-specific modeling roadmap that previously constrained future changes.",
            "Deleting the NOx/PM2.5 modeling framework further reduces the amount of explicit permit text governing when and how future modeling must be refreshed. If MPCA expects Appendix B, reopening authority, or case-by-case review to replace those detailed requirements, that expectation is not stated with the same clarity in the draft.",
            "No immediate operational effect, but future projects affecting NOx or PM2.5 emissions could face more permitting uncertainty and a less predictable review path.",
            "High",
            "Flag for client discussion and confirm whether MPCA intentionally removed the NOx/PM2.5 EBD framework or expects it to be preserved elsewhere.",
            "Significant changes",
            6,
        ),
        RowSpec(
            "PC-006",
            "No direct equivalent",
            "5.1.57-5.1.58",
            "Ambient Conditions",
            "Administrative",
            manual("No comparable General Public Preclusion Plan requirements appear in the issued permit in this section."),
            new_req_group(["5.1.57", "5.1.58"]),
            "Addition",
            "The draft adds a General Public Preclusion Plan framework built around an effective fence line and a required boundary map in Appendix C. It also prescribes minimum plan contents such as access-point identification, fencing, patrol locations, signage spacing, and a breach response plan. This is a material new compliance framework rather than a wording cleanup.",
            "These provisions affect how the source demonstrates non-ambient air and supports NAAQS compliance assumptions at the property boundary. They create new enforceable planning and documentation obligations tied to ambient-boundary control.",
            "The facility may need physical controls, signage placement, and internal procedures that were not spelled out in the prior permit language.",
            "High",
            "Confirm with agency that Appendix C and the plan requirements reflect existing site controls and do not require unplanned capital or staffing changes.",
            "Significant changes",
            3,
        ),
        RowSpec(
            "PC-007",
            "No direct equivalent",
            "5.1.59",
            "Reporting",
            "Deviations",
            manual("No comparable ambient-boundary breach documentation and reporting requirement appears in the issued permit."),
            new_req_group(["5.1.59"]),
            "Addition",
            "The draft creates a new requirement to document ambient-boundary breaches within three calendar days and report the breach and corrective measures in the semiannual deviations report. It also defines both direct and indirect evidence of a breach. That is a concrete new recordkeeping and reporting obligation with enforcement implications.",
            "This adds a new deviation-reporting pathway tied to non-ambient air boundary control rather than emissions monitoring. Failure to document a breach or corrective action could itself become a compliance issue.",
            "The facility will need a process for identifying, documenting, and escalating access breaches quickly enough to meet the three-day recordkeeping deadline.",
            "High",
            "Check consistency with current site security practices and flag for client discussion if the plant lacks a reliable mechanism to detect indirect breaches.",
            "Significant changes",
            7,
        ),
        RowSpec(
            "PC-008",
            "No direct equivalent",
            "5.1.60-5.1.61",
            "Monitoring",
            "Recordkeeping",
            manual("No comparable fence-inspection and gated-access camera requirement appears in the issued permit."),
            new_req_group(["5.1.60", "5.1.61"]),
            "Addition",
            "The draft requires fencing along the effective fence line, quarterly fence inspections with maintenance records, and controlled access at roads, trails, and abandoned rail grades using barriers and monitored or locked gates. It also requires cameras at the roadway entrance. These are operationally meaningful additions even though they are not emissions-limit changes.",
            "The new terms expand enforceable monitoring and recordkeeping obligations tied to ambient-boundary control. They also create a clearer basis for MPCA to assess whether non-ambient air assumptions remain valid.",
            "The facility may need new inspections, camera maintenance, and potentially physical access-control upgrades.",
            "Medium",
            "Confirm whether existing fencing and gate controls already satisfy the draft language or whether additional site work is needed.",
            "Significant changes",
            12,
        ),
        RowSpec(
            "PC-009",
            "No direct equivalent",
            "5.1.62-5.1.63",
            "Monitoring",
            "Administrative",
            manual("No comparable daily patrol and daily remote-monitoring obligation appears in the issued permit."),
            new_req_group(["5.1.62", "5.1.63"]),
            "Addition",
            "The draft adds a formal security patrol plan with at least daily patrols for designated effective-fence-line segments and separate daily remote monitoring at the roadway entrance. That goes beyond a general boundary concept and creates recurring operational tasks with associated records. The change is substantive because it adds ongoing compliance work rather than occasional planning documentation.",
            "These provisions increase the monitoring burden and create additional records that may be reviewed during inspections or deviations reporting. Failure to patrol or review monitoring equipment daily could become a permit deviation if the final language is retained.",
            "The plant may need staff time, contractor support, or procedural changes to keep up with patrol frequency and monitoring review expectations.",
            "Medium",
            "Check consistency with current operations and confirm whether this reflects existing practice before the draft is finalized.",
            "Significant changes",
            11,
        ),
        RowSpec(
            "PC-010",
            "No direct equivalent",
            "5.1.64-5.1.65",
            "Fugitive Emissions",
            "General Conditions",
            manual("No direct equivalent spill-cleanup RACT clause or public-nuisance clause appears in this location in the issued permit."),
            new_req_group(["5.1.64", "5.1.65"]),
            "Addition",
            "The draft adds a requirement to clean up spilled commodities as needed to minimize fugitive emissions and separately adds a public-nuisance clause for commodity-facility operation. The spill-cleanup language is a manageable housekeeping requirement, but the nuisance clause introduces broader discretionary language. The two items are related operationally, but the nuisance term carries more interpretive risk than the cleanup term alone.",
            "The spill-cleanup term reinforces fugitive dust control expectations. The nuisance language could broaden the basis for agency enforcement if off-site impacts arise even where a numeric permit limit is not clearly exceeded.",
            "The practical effect is more attention to housekeeping, truck routes, and complaint response. The nuisance clause may also influence how aggressively the facility responds to community complaints.",
            "Medium",
            "Confirm with agency whether the nuisance language is standard for this permit form and consider comment if the client wants narrower wording.",
            "Significant changes",
            13,
        ),
    ]
    rows.extend(
        [
            RowSpec(
                "PC-011",
                "5.46.1 and 5.57.1",
                "5.46.1 and 5.57.1",
                "Control Equipment",
                "Operational Limits",
                old_req_group(["5.46.1", "5.57.1"]),
                new_req_group(["5.46.1", "5.57.1"]),
                "Revision",
                "The draft keeps the normal requirement to route STRU 11 emissions through TREA 2, but it now identifies the specific EQUIs in AOS 1 and revises the parallel section 5.57.1 control language so it applies only when operating under AOS 1. This is mostly a clarification of the normal operating case, but it also sets up the draft's new bypass scenario by carving out when the control requirement applies. The real compliance significance depends on how AOS 2 is used.",
                "On its own, the revision makes the normal controlled operating scenario more explicit. It also narrows the scope of the blanket control requirement by tying it to AOS 1 rather than all STRU 11 operation.",
                "Little change to normal operation so long as the plant continues routing STRU 11 emissions to TREA 2 outside of maintenance events.",
                "Medium",
                "Verify the listed EQUIs match the application and current equipment inventory.",
                "Minor changes, likely acceptable",
                18,
            ),
            RowSpec(
                "PC-012",
                "No direct equivalent",
                "5.46.2",
                "Control Equipment",
                "Operational Limits",
                manual("The issued permit does not authorize an alternate operating scenario that bypasses TREA 2 for STRU 11 maintenance."),
                new_req_group(["5.46.2"]),
                "Addition",
                "The draft adds AOS 2 for STRU 11 and allows EQUIs 18, 191, and 243 to vent directly to atmosphere while TREA 2 is undergoing maintenance. That is one of the most substantive changes in the draft because it authorizes a direct-vent operating mode that the issued permit did not clearly allow. Even though the scenario is limited to maintenance, it changes both control strategy and compliance posture.",
                "This is a new permit authorization affecting when control equipment must operate and when bypassed emissions are permissible. It should be treated as a material change in permit scope and enforceability.",
                "The facility gains maintenance flexibility, but it must manage the bypass carefully because the scenario will likely draw scrutiny from both the agency and client stakeholders.",
                "High",
                "Flag for client discussion and verify the application, emissions calculations, and modeled assumptions fully support this bypass scenario.",
                "Significant changes",
                1,
            ),
            RowSpec(
                "PC-013",
                "5.46.2-5.46.7",
                "5.46.3-5.46.8",
                "Emission Limits",
                "Opacity",
                old_req_group(["5.46.2", "5.46.3", "5.46.4", "5.46.5", "5.46.6", "5.46.7"]),
                new_req_group(["5.46.3", "5.46.4", "5.46.5", "5.46.6", "5.46.7", "5.46.8"]),
                "Revision",
                "The numerical PM, PM10, PM2.5, and opacity limits are carried forward, but the draft now states these limits apply when operating in AOS 1. The AOS 2 block does not provide a parallel set of numeric emission or opacity limits for the direct-vent scenario. That creates a meaningful interpretive issue even if the hour cap and supporting emissions analysis are intended to control the bypass case.",
                "The draft narrows the explicit applicability of existing emission-limit language and relies more heavily on scenario-specific operational restrictions for the maintenance bypass case. That can create ambiguity over what enforceable emission benchmarks apply during AOS 2.",
                "Operationally, the plant may be able to use the bypass more flexibly, but compliance demonstrations during AOS 2 could become harder to explain if the final permit does not state the governing emission expectations clearly.",
                "High",
                "Consider comment on draft and ask the agency to clarify what emission-limit framework applies during AOS 2.",
                "Significant changes",
                2,
            ),
            RowSpec(
                "PC-014",
                "No direct equivalent",
                "5.46.9-5.46.11",
                "Operational Limits",
                "Recordkeeping",
                manual("The issued permit does not contain a 100-hour annual AOS 2 cap or the corresponding AOS 2 hour-tracking requirements."),
                new_req_group(["5.46.9", "5.46.10", "5.46.11"]),
                "Addition",
                "The draft adds a hard cap of 100 hours per year for operation under AOS 2 and adds daily and monthly recordkeeping to track those hours. That is a substantive new operating restriction paired with new records that will be critical to proving the bypass remains within the permitted scope. The added records are likely the main control on direct-vent duration.",
                "This creates a new enforceable annual operating limit plus daily and monthly records that can independently support deviation findings. The cap is likely central to the source's NSR and modeling basis for the bypass scenario.",
                "The facility will need reliable outage logging and monthly rolling calculations whenever TREA 2 maintenance occurs.",
                "High",
                "Check consistency with current operations and verify the client can track AOS 2 hours cleanly enough to defend the 12-month rolling total.",
                "Significant changes",
                4,
            ),
            RowSpec(
                "PC-015",
                "5.46.10",
                "5.46.14",
                "Throughput or Production Limits",
                "Operational Limits",
                old_req_group(["5.46.10"]),
                new_req_group(["5.46.14"]),
                "Revision",
                "The STRU 11 short-term DDGS loadout limit drops from 199.8 tons per hour to 178.8 tons per hour, and the draft ties the revised value to the October 30, 2025 performance test rather than the November 20, 2023 test. This is a direct change to the permitted operating envelope. It appears test-supported, but it still tightens the allowed short-term rate.",
                "The change reduces the permitted throughput benchmark that supports compliance demonstrations and future reset calculations for STRU 11. It is therefore a real substantive change even though the surrounding recordkeeping language is largely the same.",
                "This could constrain DDGS loadout scheduling and maintenance planning if operations previously relied on the higher 199.8 tph limit.",
                "High",
                "Verify basis in application and check consistency with current operations before accepting the lower throughput value.",
                "Significant changes",
                8,
            ),
            RowSpec(
                "PC-016",
                "5.46.8-5.46.17",
                "5.46.12-5.46.21",
                "Operational Limits",
                "Performance Testing",
                old_req_group(["5.46.8", "5.46.9", "5.46.11", "5.46.12", "5.46.13", "5.46.14", "5.46.15", "5.46.16", "5.46.17"]),
                new_req_group(["5.46.12", "5.46.13", "5.46.15", "5.46.16", "5.46.17", "5.46.18", "5.46.19", "5.46.20", "5.46.21"]),
                "Possible relocation or renumbering",
                "Most of the existing STRU 11 maximum-achievable-rate, reset-protocol, equipment-flexibility, amendment-trigger, and performance-testing-recordkeeping language is retained but shifted to higher requirement numbers after the new AOS provisions are inserted. Aside from the new throughput value captured separately, this block mostly reads as a relocation and renumbering exercise. The practical compliance structure remains intact.",
                "Because the duties remain substantially the same, this appears to be a structural renumbering rather than a separate new burden. The main compliance effect is citation cleanup for internal cross-references and procedures.",
                "No material operational change beyond the separate throughput revision already noted.",
                "Low",
                "No action needed other than updating any internal permit cross-references that cite the old requirement numbers.",
                "No change",
                24,
            ),
            RowSpec(
                "PC-017",
                "6.10.1-6.10.3",
                "6.10.1-6.10.3",
                "Performance Testing",
                "Operational Limits",
                old_req_group(["6.10.1", "6.10.2", "6.10.3"]),
                new_req_group(["6.10.1", "6.10.2", "6.10.3"]),
                "Revision",
                "The periodic STRU 11 PM, PM2.5, and particulate-matter testing rows are now labeled as AOS 1 tests in the draft. The core due dates and methods remain the same, so the change reads mainly as scenario clarification rather than a new testing burden. Still, it is worth noting because it reinforces that the periodic stack tests are tied to the controlled operating scenario, not the new maintenance bypass case.",
                "This helps clarify which operating mode the routine STRU 11 tests are intended to represent. It does not appear to add or remove a test, but it narrows interpretive ambiguity.",
                "No major operational change beyond clearer separation between normal operation and the maintenance-bypass scenario.",
                "Low",
                "No action needed unless the client wants the permit to state explicitly how AOS 2 emissions will be verified.",
                "Minor changes, likely acceptable",
                23,
            ),
            RowSpec(
                "PC-018",
                "6.9.4, 6.10.4, 6.11.4",
                "6.9.4, 6.10.4, 6.11.4",
                "Performance Testing",
                "Testing",
                old_req_group(["6.9.4", "6.10.4", "6.11.4"]),
                new_req_group(["6.9.4", "6.10.4", "6.11.4"]),
                "Revision",
                "For the post-modification particulate-matter tests at STRU 10, STRU 11, and STRU 12, the draft uses EPA Methods 5 and 202 instead of the issued permit's Method 201A and 202 combination in the particulate-matter rows. That looks like a technical correction because Method 5 is the more typical particulate method for total PM. It is still a real testing-method revision and should not be dismissed as formatting.",
                "The test method used to demonstrate compliance is an enforceable element of the permit. Even if the revision is likely appropriate, it changes the compliance protocol that would apply after equipment changes.",
                "This should not materially affect routine operation, but it could change planning and contractor scope for future post-modification testing events.",
                "Medium",
                "Verify basis in application or agency testing rationale and confirm the revised test methods are intentional.",
                "Minor changes, likely acceptable",
                15,
            ),
            RowSpec(
                "PC-019",
                "5.48.16",
                "5.48.16",
                "Throughput or Production Limits",
                "Operational Limits",
                old_req_group(["5.48.16"]),
                new_req_group(["5.48.16"]),
                "Revision",
                "The short-term beer-feed limit at STRU 13 decreases slightly from 1,974 gpm to 1,968.8 gpm, and the draft updates the supporting test reference from April 2023 to October 30, 2025. The numerical change is small, but it is still a tightening of the permitted short-term operating rate. This is best viewed as a modest substantive update driven by new test data.",
                "The revised value affects the operating benchmark used for compliance and for future throughput-reset logic. Because the reduction is small, the compliance impact is likely manageable unless operations were already close to the old cap.",
                "Minimal operational effect is expected, but the lower value should be checked against current operating practice and control margins.",
                "Medium",
                "Verify basis in application and confirm the small reduction will not constrain current production planning.",
                "Minor changes, likely acceptable",
                16,
            ),
            RowSpec(
                "PC-020",
                "6.12.1-6.12.4",
                "6.12.1-6.12.4",
                "Performance Testing",
                "Reporting",
                old_req_group(["6.12.1", "6.12.2", "6.12.3", "6.12.4"]),
                new_req_group(["6.12.1", "6.12.2", "6.12.3", "6.12.4"]),
                "Revision",
                "The periodic STRU 13 VOC and HAP tests move from a due date of April 13, 2026 to October 30, 2028 for both AOS 1 and AOS 2. The recurring interval remains the same, so the draft appears to be resetting the schedule based on more recent performance testing. This is a substantive but likely manageable reduction in near-term testing burden.",
                "The change delays the next required round of STRU 13 periodic testing. That reduces immediate compliance activity, but it should be supported by completed testing and agency acceptance of the newer baseline date.",
                "Operationally this reduces near-term stack-testing burden and associated scheduling pressure.",
                "Medium",
                "Confirm whether the October 30, 2025 test has already been accepted by MPCA and supports the revised due dates.",
                "Minor changes, likely acceptable",
                14,
            ),
            RowSpec(
                "PC-021",
                "5.48.11-5.48.12",
                "5.48.11-5.48.12",
                "Other",
                "Possible relocation or renumbering",
                old_req_group(["5.48.11", "5.48.12"]),
                new_req_group(["5.48.11", "5.48.12"]),
                "Possible relocation or renumbering",
                "The methanol and hexane AOS 2 controlled emission-factor rows switch order in the draft. The actual numerical values and function of the two factors do not appear to change. This reads as a renumbering or ordering cleanup rather than a substantive revision.",
                "No apparent change to how emission factors are used for compliance calculations. The main effect is on citation order only.",
                "No operational impact.",
                "Low",
                "No action needed.",
                "No change",
                26,
            ),
            RowSpec(
                "PC-022",
                "5.59.13",
                "5.59.13",
                "CAM",
                "Monitoring",
                old_req_group(["5.59.13"]),
                new_req_group(["5.59.13"]),
                "Revision",
                "The CO2 scrubber AOS 1 fresh-water CAM limit falls from 111.0 gpm to 100.0 gpm, and the supporting test reference updates to October 30, 2025. That makes the indicator less stringent than the issued permit. The change appears test-driven, but it still changes the compliance threshold used to determine when emissions are considered uncontrolled.",
                "A lower indicator threshold reduces the control margin before a deviation occurs. If supported by recent testing, it may be acceptable, but it still alters the CAM basis and deviation trigger.",
                "This may reduce water-demand burden slightly during operation, but it also changes how the control system is expected to perform under AOS 1.",
                "Medium",
                "Verify basis in application and confirm the revised CAM threshold is supported by accepted test results.",
                "Minor changes, likely acceptable",
                17,
            ),
            RowSpec(
                "PC-023",
                "5.59.15",
                "5.59.15",
                "CAM",
                "Monitoring",
                old_req_group(["5.59.15"]),
                new_req_group(["5.59.15"]),
                "Revision",
                "The recirculating scrubber water CAM limit for AOS 1 rises from 125.0 gpm to 145.0 gpm based on the October 30, 2025 performance test. Unlike the fresh-water change, this is a tightening of the monitored parameter. That is a meaningful operational and compliance revision because excursions will occur sooner if the recirculating rate drops.",
                "The higher CAM threshold increases the monitoring burden and tightens the margin for avoiding deviations. Because the permit treats periods below the threshold as uncontrolled, this change can directly affect deviation reporting and compliance status.",
                "The facility may need to maintain higher recirculation rates or adjust operating practices to stay above the tighter indicator value.",
                "High",
                "Check consistency with current operations and confirm equipment can reliably sustain the higher recirculation rate.",
                "Significant changes",
                9,
            ),
            RowSpec(
                "PC-024",
                "5.59.27-5.59.28",
                "5.59.27-5.59.28",
                "CAM",
                "Recordkeeping",
                old_req_group(["5.59.27", "5.59.28"]),
                new_req_group(["5.59.27", "5.59.28"]),
                "Possible relocation or renumbering",
                "The issued permit places recirculating-scrubber-water recordkeeping at 5.59.27 and ABS additive recordkeeping at 5.59.28, while the draft flips that order. The underlying duties appear unchanged. This looks like a numbering or ordering correction only.",
                "No meaningful change to CAM recordkeeping content is evident. The effect is limited to citation cleanup.",
                "No operational impact.",
                "Low",
                "No action needed.",
                "No change",
                25,
            ),
            RowSpec(
                "PC-025",
                "5.61.24-5.61.29",
                "5.61.24-5.61.29",
                "CAM",
                "Monitoring",
                old_req_group(["5.61.24", "5.61.25", "5.61.28", "5.61.29"]),
                new_req_group(["5.61.24", "5.61.25", "5.61.28", "5.61.29"]),
                "Possible relocation or renumbering",
                "The thermal oxidizer CAM provisions appear to be reordered, with monitoring-equipment and temperature-monitoring text switching sequence and the annual-calibration and accumulation-of-deviations items likewise moving. The substantive duties themselves do not appear to change. This looks like a drafting cleanup rather than a revised CAM program.",
                "No material compliance change is apparent as long as the final wording stays the same. Internal cross-references may need updating if any procedures cite the old numbering order.",
                "No operational change.",
                "Low",
                "No action needed other than checking internal citations if the client tracks requirement numbers closely.",
                "No change",
                27,
            ),
            RowSpec(
                "PC-026",
                "6.13.4",
                "6.13.4",
                "Performance Testing",
                "Fuel Requirements",
                old_req_group(["6.13.4"]),
                new_req_group(["6.13.4"]),
                "Revision",
                "The periodic STRU 14 sulfur-dioxide testing requirement is materially simplified in the draft. The issued permit let the Commissioner set a tighter future test frequency by NOC/NOV letter and required the periodic test to be run with dryer-feed sulfur content below 2.0 percent dry basis with sulfur-content information in the test plan. The draft keeps the 60-month schedule and Method 6C but drops those additional constraints.",
                "This removes explicit permit language that could have increased periodic SO2 testing frequency and removes a specific low-sulfur testing condition from the routine periodic test requirement. That is a real change to the enforceable testing framework, even though it may reduce burden.",
                "The plant may gain flexibility in how it runs the periodic SO2 test and may avoid a more aggressive future retest schedule unless the permit is revised again.",
                "Medium",
                "Confirm with agency whether the deleted frequency-setting and sulfur-content language was intentional and whether those expectations now appear elsewhere outside the permit text.",
                "Significant changes",
                10,
            ),
            RowSpec(
                "PC-027",
                "5.1.1",
                "5.1.1",
                "Temporary or Construction Conditions",
                "Administrative",
                old_req_group(["5.1.1"]),
                new_req_group(["5.1.1"]),
                "Revision",
                "The recycling-stream construction authorization is carried forward, but the draft adds the explicit parenthetical date '(December 12, 2027)' to the statement that the authorization expires five years after issuance of Permit 09100062-101. That clarifies the deadline rather than changing the deadline itself. The revision is useful because it removes the need to back-calculate the date.",
                "This is primarily a clarification of an existing construction-expiration condition. It should reduce interpretation disputes about the expiration date.",
                "No change to facility operation if the client already understood the underlying five-year expiration date.",
                "Low",
                "No action needed.",
                "Minor changes, likely acceptable",
                22,
            ),
        ]
    )
    return rows


def row_to_values(row: RowSpec) -> list[str]:
    request_status, source, basis, priority = request_meta(row.comparison_id)
    validation_basis, evidence_confidence = summarize_validation(row)
    return [
        row.comparison_id,
        row.old_section,
        row.new_section,
        row.subject_primary,
        row.subject_secondary,
        row.old_text_getter(),
        row.new_text_getter(),
        row.change_type,
        row.initial_analysis,
        row.regulatory_impact,
        row.operational_impact,
        row.risk_level,
        row.follow_up,
        row.conclusion,
        request_status,
        source,
        basis,
        priority,
        validation_basis,
        evidence_confidence,
    ]


def write_permit_text_sheet(ws, reqs: OrderedDict[str, str], header_fill, header_font, wrap) -> None:
    ws.append(["Requirement Number", "Condition Text"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for req_id, text in reqs.items():
        ws.append([req_id, f"{req_id} {text}".strip()])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    autofit(ws, {1: 18, 2: 135})


def write_matched_conditions_sheet(ws, header_fill, header_font, wrap) -> None:
    headers = ["Match ID", "Match Status", "Old Permit Section", "New Permit Section", "Similarity", "Old Text", "New Text"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for row in build_matched_conditions():
        ws.append([row[h] for h in headers])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    autofit(ws, {1: 12, 2: 16, 3: 18, 4: 18, 5: 10, 6: 95, 7: 95})


def write_application_basis_sheet(ws, rows: list[RowSpec], header_fill, header_font, wrap) -> None:
    ws["A1"] = "Application File"
    ws["B1"] = "Current Relevance"
    ws["C1"] = "Key Requested Change Summary"
    ws["D1"] = "Analytical Use"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    line = 2
    for name, relevance, summary, note in APPLICATION_FILES:
        ws.append([name, relevance, summary, note])
        line += 1

    line += 1
    ws[f"A{line}"] = "Comparison ID"
    ws[f"B{line}"] = "Requested Change Status"
    ws[f"C{line}"] = "Application Source"
    ws[f"D{line}"] = "Application Basis"
    ws[f"E{line}"] = "Review Priority"
    for cell in ws[line]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    line += 1
    for row in rows:
        request_status, source, basis, priority = request_meta(row.comparison_id)
        ws[f"A{line}"] = row.comparison_id
        ws[f"B{line}"] = request_status
        ws[f"C{line}"] = source
        ws[f"D{line}"] = basis
        ws[f"E{line}"] = priority
        line += 1
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    autofit(ws, {1: 22, 2: 22, 3: 56, 4: 72, 5: 14})


def write_qa_spot_checks_sheet(ws, header_fill, header_font, wrap) -> None:
    headers = [
        "QA ID",
        "Workflow Sheet",
        "Subject Item",
        "Requirement / Topic",
        "Workflow Claim",
        "Manual Permit Check Result",
        "Old Permit Evidence",
        "Draft Permit Evidence",
        "QA Conclusion",
        "Recommended Use",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    for finding in MANUAL_QA_FINDINGS:
        ws.append([finding[h] for h in headers])

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    autofit(ws, {1: 10, 2: 20, 3: 16, 4: 18, 5: 46, 6: 52, 7: 40, 8: 40, 9: 24, 10: 42})


def write_validation_matrix_sheet(ws, rows: list[RowSpec], header_fill, header_font, wrap) -> None:
    headers = [
        "Comparison ID",
        "Old Permit Section",
        "New Permit Section",
        "Old Evidence Reqs",
        "New Evidence Reqs",
        "Old Req Count",
        "New Req Count",
        "Validation Basis",
        "Evidence Confidence",
        "Consultant Conclusion",
        "Review Priority",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    for row in rows:
        old_refs = expand_section_reference(row.old_section, OLD_REQS)
        new_refs = expand_section_reference(row.new_section, NEW_REQS)
        validation_basis, evidence_confidence = summarize_validation(row)
        ws.append(
            [
                row.comparison_id,
                row.old_section,
                row.new_section,
                ", ".join(old_refs),
                ", ".join(new_refs),
                len(old_refs),
                len(new_refs),
                validation_basis,
                evidence_confidence,
                row.conclusion,
                request_meta(row.comparison_id)[3],
            ]
        )

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    autofit(ws, {1: 14, 2: 24, 3: 24, 4: 34, 5: 34, 6: 10, 7: 10, 8: 42, 9: 18, 10: 26, 11: 14})


WORKFLOW_ALIGNMENT_MAP = {
    ("Requirement Comparison", "5.1.59"): ("Aligned with consultant analysis", "PC-007", "Captured within the grouped ambient-boundary breach reporting analysis."),
    ("Requirement Comparison", "5.1.60"): ("Aligned with consultant analysis", "PC-008", "Captured within the grouped fencing and access-control analysis."),
    ("Requirement Comparison", "5.1.61"): ("Aligned with consultant analysis", "PC-008", "Captured within the grouped fencing and access-control analysis."),
    ("Requirement Comparison", "5.1.63"): ("Aligned with consultant analysis", "PC-009", "Captured within the grouped patrol and remote-monitoring analysis."),
    ("Requirement Comparison", "5.46.14"): ("Aligned with consultant analysis", "PC-015", "Captured as the tightened STRU 11 throughput limit."),
    ("Requirement Comparison", "5.48.16"): ("Aligned with consultant analysis", "PC-019", "Captured as the STRU 13 throughput revision."),
    ("Requirement Comparison", "5.59.15"): ("Aligned with consultant analysis", "PC-023", "Captured as the tighter recirculating scrubber-water CAM threshold."),
    ("Requirement Comparison", "6.12.4"): ("Aligned with consultant analysis", "PC-020", "Captured within the STRU 13 periodic testing schedule update."),
    ("Removed & Added", "5.1.57"): ("Aligned with consultant analysis", "PC-004", "Grouped with the deleted PM10 modeling trigger framework."),
    ("Removed & Added", "5.1.58"): ("Aligned with consultant analysis", "PC-004", "Grouped with the deleted PM10 modeling trigger framework."),
    ("Removed & Added", "5.1.59"): ("Aligned with consultant analysis", "PC-005", "Grouped with the deleted PM10 modeling protocol framework."),
    ("Removed & Added", "5.1.60"): ("Aligned with consultant analysis", "PC-005", "Grouped with the deleted PM10 modeling protocol framework."),
    ("Removed & Added", "5.1.61"): ("Aligned with consultant analysis", "PC-005", "Grouped with the deleted PM10 modeling protocol framework."),
    ("Removed & Added", "5.1.62"): ("Aligned with consultant analysis", "PC-005", "Grouped with the deleted PM10 modeling protocol framework."),
    ("Removed & Added", "5.1.63"): ("Aligned with consultant analysis", "PC-005", "Grouped with the deleted PM10 refined-modeling trigger framework."),
    ("Removed & Added", "5.1.64"): ("Aligned with consultant analysis", "PC-005", "Grouped with the deleted PM10 protocol deadline framework."),
    ("Removed & Added", "5.1.65"): ("Aligned with consultant analysis", "PC-005", "Grouped with the deleted PM10 protocol deadline framework."),
    ("Removed & Added", "5.1.66"): ("Aligned with consultant analysis", "PC-005", "Grouped with the deleted PM10 final-modeling-report requirement."),
    ("Removed & Added", "5.1.67"): ("Additional meaningful issue found in cross-check", "PC-028", "The workflow flagged the deletion, and the consultant analysis now separately captures the NOx/PM2.5 modeling framework removal."),
    ("Removed & Added", "5.1.68"): ("Additional meaningful issue found in cross-check", "PC-028", "The workflow flagged the deletion, and the consultant analysis now separately captures the NOx/PM2.5 modeling framework removal."),
    ("Removed & Added", "5.1.69"): ("Additional meaningful issue found in cross-check", "PC-028", "The workflow flagged the deletion, and the consultant analysis now separately captures the NOx/PM2.5 modeling framework removal."),
    ("Removed & Added", "5.1.70"): ("Additional meaningful issue found in cross-check", "PC-028", "The workflow flagged the deletion, and the consultant analysis now separately captures the NOx/PM2.5 modeling framework removal."),
    ("Removed & Added", "5.1.71"): ("Additional meaningful issue found in cross-check", "PC-028", "The workflow flagged the deletion, and the consultant analysis now separately captures the NOx/PM2.5 modeling framework removal."),
    ("Removed & Added", "5.1.72"): ("Additional meaningful issue found in cross-check", "PC-028", "The workflow flagged the deletion, and the consultant analysis now separately captures the NOx/PM2.5 modeling framework removal."),
    ("Removed & Added", "5.1.73"): ("Additional meaningful issue found in cross-check", "PC-028", "The workflow flagged the deletion, and the consultant analysis now separately captures the NOx/PM2.5 modeling framework removal."),
    ("Removed & Added", "5.1.74"): ("Additional meaningful issue found in cross-check", "PC-028", "The workflow flagged the deletion, and the consultant analysis now separately captures the NOx/PM2.5 modeling framework removal."),
    ("Removed & Added", "5.1.75"): ("Additional meaningful issue found in cross-check", "PC-028", "The workflow flagged the deletion, and the consultant analysis now separately captures the NOx/PM2.5 modeling framework removal."),
    ("Removed & Added", "5.1.76"): ("Additional meaningful issue found in cross-check", "PC-028", "The workflow flagged the deletion, and the consultant analysis now separately captures the NOx/PM2.5 modeling framework removal."),
    ("Removed & Added", "5.1.77"): ("Additional meaningful issue found in cross-check", "PC-028", "The workflow flagged the deletion, and the consultant analysis now separately captures the NOx/PM2.5 modeling framework removal."),
    ("Removed & Added", "5.46.9"): ("Aligned with consultant analysis", "PC-014", "Captured within the grouped AOS 2 hour-cap and recordkeeping additions."),
    ("Removed & Added", "5.46.10"): ("Aligned with consultant analysis", "PC-014", "Captured within the grouped AOS 2 hour-cap and recordkeeping additions."),
    ("Removed & Added", "5.46.11"): ("Aligned with consultant analysis", "PC-014", "Captured within the grouped AOS 2 hour-cap and recordkeeping additions."),
}


def load_workflow_items() -> list[dict[str, str]]:
    if not WORKFLOW_XLSX.exists():
        return []
    wb = load_workbook(WORKFLOW_XLSX, read_only=True, data_only=True)
    items: list[dict[str, str]] = []

    ws = wb["Requirement Comparison"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[1]
    for row in rows[2:]:
        if not any(v is not None for v in row):
            continue
        item = dict(zip(headers, row))
        if item.get("Impact") not in {"MEDIUM", "HIGH"}:
            continue
        items.append(
            {
                "Workflow Sheet": "Requirement Comparison",
                "Workflow Mode": item.get("Change Type", ""),
                "Subject Item": item.get("Subject Item", ""),
                "Requirement": item.get("Req #", ""),
                "Impact": item.get("Impact", ""),
                "Workflow Summary": item.get("Change Summary", ""),
                "Old Text": item.get("Text", "") or "",
                "New Text": item.get("Text", "") or "",
            }
        )

    ws = wb["Removed & Added"]
    rows = list(ws.iter_rows(values_only=True))
    mode = ""
    headers = None
    for row in rows:
        if row[0] == "REMOVED REQUIREMENTS":
            mode = "Removed"
            headers = None
            continue
        if row[0] == "ADDED REQUIREMENTS":
            mode = "Added"
            headers = None
            continue
        if row[0] == "Subject Item":
            headers = row
            continue
        if not headers or not any(v is not None for v in row):
            continue
        item = dict(zip(headers, row))
        if item.get("Impact") not in {"MEDIUM", "HIGH"}:
            continue
        text_key = "Original Text" if mode == "Removed" else "Draft Text"
        items.append(
            {
                "Workflow Sheet": "Removed & Added",
                "Workflow Mode": mode,
                "Subject Item": item.get("Subject Item", ""),
                "Requirement": item.get("Requirement", ""),
                "Impact": item.get("Impact", ""),
                "Workflow Summary": item.get("Reason", ""),
                "Old Text": item.get(text_key, "") or "",
                "New Text": item.get(text_key, "") or "",
            }
        )

    ws = wb["Equipment Summaries"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        item = dict(zip(headers, row))
        if item.get("Overall Impact") not in {"MEDIUM", "HIGH"}:
            continue
        items.append(
            {
                "Workflow Sheet": "Equipment Summaries",
                "Workflow Mode": "Summary",
                "Subject Item": item.get("Subject Item", "") or "",
                "Requirement": item.get("Subject Item", "") or "",
                "Impact": item.get("Overall Impact", "") or "",
                "Workflow Summary": item.get("Change Summary", "") or "",
                "Old Text": item.get("Key Changes", "") or "",
                "New Text": item.get("Flags", "") or "",
            }
        )

    wb.close()
    return items


def best_requirement_match(text: str, candidates: OrderedDict[str, str]) -> tuple[str, float]:
    best_req = ""
    best_score = 0.0
    for req_id, req_text in candidates.items():
        score = condition_similarity(text, req_text)
        if score > best_score:
            best_req = req_id
            best_score = score
    return best_req, best_score


def classify_workflow_item(item: dict[str, str]) -> tuple[str, str, str]:
    key = (item["Workflow Sheet"], item["Requirement"])
    if key in WORKFLOW_ALIGNMENT_MAP:
        return WORKFLOW_ALIGNMENT_MAP[key]

    req = item["Requirement"]
    if item["Workflow Sheet"] == "Equipment Summaries":
        if item["Subject Item"] == "COMG 2":
            return (
                "Workflow summary overcall",
                "",
                "Manual QA found that the COMG 2 summary overstates removals. Draft requirements 5.2.9, 5.2.10, and 5.2.11 remain in place.",
            )
        return (
            "Needs manual comparison",
            "",
            "Equipment-level summary rows are useful screening notes, but they were not used as standalone consultant conclusions without permit-text validation.",
        )

    if item["Workflow Sheet"] == "Removed & Added":
        if item["Workflow Mode"] == "Removed":
            if req in NEW_REQS:
                return (
                    "Workflow overcall / equivalent retained",
                    "",
                    f"The draft still contains requirement {req}. The workflow workbook appears to have treated a revised condition as removed.",
                )
            best_req, best_score = best_requirement_match(item["Old Text"], NEW_REQS)
            if best_score >= 0.78:
                return (
                    "Workflow overcall / likely renumbered",
                    "",
                    f"The old requirement appears to remain in the draft in equivalent form at {best_req} (similarity {best_score:.2f}).",
                )
        if item["Workflow Mode"] == "Added":
            if req in OLD_REQS:
                return (
                    "Workflow overcall / already in issued permit",
                    "",
                    f"The issued permit already contains requirement {req}; the workflow workbook appears to have treated it as newly added.",
                )
            best_req, best_score = best_requirement_match(item["New Text"], OLD_REQS)
            if best_score >= 0.78:
                return (
                    "Workflow overcall / likely existing equivalent",
                    "",
                    f"The draft requirement appears materially similar to issued-permit requirement {best_req} (similarity {best_score:.2f}).",
                )

    return ("Needs manual comparison", "", "This workflow flag was not automatically matched to the consultant analysis or to an obvious renumbered/equivalent requirement.")


def write_workflow_crosscheck_sheet(ws, header_fill, header_font, wrap) -> None:
    items = load_workflow_items()
    headers = [
        "Workflow Sheet",
        "Workflow Mode",
        "Subject Item",
        "Requirement",
        "Impact",
        "Workflow Summary",
        "Consultant Cross-Check",
        "Consultant Row ID",
        "Cross-Check Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    for item in items:
        cross_check, consultant_row, note = classify_workflow_item(item)
        ws.append(
            [
                item["Workflow Sheet"],
                item["Workflow Mode"],
                item["Subject Item"],
                item["Requirement"],
                item["Impact"],
                item["Workflow Summary"],
                cross_check,
                consultant_row,
                note,
            ]
        )

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    autofit(ws, {1: 22, 2: 14, 3: 16, 4: 14, 5: 10, 6: 72, 7: 34, 8: 18, 9: 72})


def write_workbook(rows: list[RowSpec], overall_outcome: str, basis_paragraph: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Comparison"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    for row in rows:
        ws.append(row_to_values(row))

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    autofit(
        ws,
        {
            1: 14,
            2: 28,
            3: 28,
            4: 22,
            5: 22,
            6: 60,
            7: 60,
            8: 22,
            9: 44,
            10: 38,
            11: 34,
            12: 12,
            13: 24,
            14: 28,
            15: 24,
            16: 34,
            17: 56,
            18: 14,
            19: 42,
            20: 18,
        },
    )

    summary = wb.create_sheet("Executive Summary")
    summary["A1"] = "Metric"
    summary["B1"] = "Value"
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    change_counts = Counter(row.change_type for row in rows)
    conclusion_counts = Counter(row.conclusion for row in rows)
    subject_counts = Counter(row.subject_primary for row in rows)
    request_counts = Counter(request_meta(row.comparison_id)[0] for row in rows)
    review_priority_counts = Counter(request_meta(row.comparison_id)[3] for row in rows)
    evidence_counts = Counter(summarize_validation(row)[1] for row in rows)
    workflow_items = load_workflow_items()
    workflow_crosscheck_counts = Counter(classify_workflow_item(item)[0] for item in workflow_items)

    line = 2
    summary[f"A{line}"] = "Total number of changes identified"
    summary[f"B{line}"] = len(rows)
    line += 2

    summary[f"A{line}"] = "Counts by Change Type"
    line += 1
    for key, value in sorted(change_counts.items()):
        summary[f"A{line}"] = key
        summary[f"B{line}"] = value
        line += 1

    line += 1
    summary[f"A{line}"] = "Counts by Consultant Conclusion"
    line += 1
    for key, value in sorted(conclusion_counts.items()):
        summary[f"A{line}"] = key
        summary[f"B{line}"] = value
        line += 1

    line += 1
    summary[f"A{line}"] = "Counts by Subject Tag Primary"
    line += 1
    for key, value in sorted(subject_counts.items()):
        summary[f"A{line}"] = key
        summary[f"B{line}"] = value
        line += 1

    line += 1
    summary[f"A{line}"] = "Counts by Requested Change Status"
    line += 1
    for key, value in sorted(request_counts.items()):
        summary[f"A{line}"] = key
        summary[f"B{line}"] = value
        line += 1

    line += 1
    summary[f"A{line}"] = "Counts by Review Priority"
    line += 1
    for key, value in sorted(review_priority_counts.items()):
        summary[f"A{line}"] = key
        summary[f"B{line}"] = value
        line += 1

    line += 1
    summary[f"A{line}"] = "Counts by Evidence Confidence"
    line += 1
    for key, value in sorted(evidence_counts.items()):
        summary[f"A{line}"] = key
        summary[f"B{line}"] = value
        line += 1

    if workflow_items:
        line += 1
        summary[f"A{line}"] = "Workflow Cross-Check Summary"
        line += 1
        for key, value in sorted(workflow_crosscheck_counts.items()):
            summary[f"A{line}"] = key
            summary[f"B{line}"] = value
            line += 1
        line += 1
        summary[f"A{line}"] = "Workflow Use Note"
        summary[f"B{line}"] = WORKFLOW_USE_NOTE
        line += 1
        summary[f"A{line}"] = "Manual QA Spot-Check Result"
        summary[f"B{line}"] = "COMG 2 workflow flags for removed 5.2.9, 5.2.10, and 5.2.11 were manually checked against the old and draft permits and were not supported by the permit text."

    top_rows = sorted(
        rows,
        key=lambda item: (
            {"Highlight": 0, "Watch": 1, "Routine": 2}[request_meta(item.comparison_id)[3]],
            item.importance_rank,
            item.comparison_id,
        ),
    )[:10]
    line += 1
    summary[f"A{line}"] = "Top 10 most important changes"
    line += 1
    summary[f"A{line}"] = "Comparison ID"
    summary[f"B{line}"] = "Summary"
    for cell in summary[line]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    line += 1
    for row in top_rows:
        summary[f"A{line}"] = row.comparison_id
        summary[f"B{line}"] = f"[{request_meta(row.comparison_id)[3]}] {row.new_section or row.old_section}: {row.initial_analysis}"
        line += 1

    line += 1
    summary[f"A{line}"] = "Overall outcome"
    summary[f"B{line}"] = overall_outcome
    line += 1
    summary[f"A{line}"] = "Basis for overall conclusion"
    summary[f"B{line}"] = basis_paragraph

    for row_cells in summary.iter_rows():
        for cell in row_cells:
            cell.alignment = wrap

    autofit(summary, {1: 34, 2: 120})

    write_permit_text_sheet(wb.create_sheet("Existing Permit Text"), OLD_REQS, header_fill, header_font, wrap)
    write_permit_text_sheet(wb.create_sheet("Draft Permit Text"), NEW_REQS, header_fill, header_font, wrap)
    write_matched_conditions_sheet(wb.create_sheet("Matched Conditions"), header_fill, header_font, wrap)
    write_application_basis_sheet(wb.create_sheet("Application Basis"), rows, header_fill, header_font, wrap)
    write_validation_matrix_sheet(wb.create_sheet("Validation Matrix"), rows, header_fill, header_font, wrap)
    write_workflow_crosscheck_sheet(wb.create_sheet("Workflow Cross-Check"), header_fill, header_font, wrap)
    write_qa_spot_checks_sheet(wb.create_sheet("QA Spot Checks"), header_fill, header_font, wrap)
    try:
        wb.save(XLSX_OUT)
        return XLSX_OUT
    except PermissionError:
        wb.save(XLSX_FALLBACK_OUT)
        return XLSX_FALLBACK_OUT


def write_memo(rows: list[RowSpec], overall_outcome: str, basis_paragraph: str) -> None:
    top_rows = sorted(
        rows,
        key=lambda item: (
            {"Highlight": 0, "Watch": 1, "Routine": 2}[request_meta(item.comparison_id)[3]],
            item.importance_rank,
            item.comparison_id,
        ),
    )[:10]
    editorial_rows = [row for row in rows if row.conclusion == "No change"]
    row_map = {row.comparison_id: row for row in rows}

    permit_wide = [row_map["PC-001"], row_map["PC-002"], row_map["PC-003"]]
    tfac_modeling = [row_map["PC-004"], row_map["PC-005"], row_map["PC-028"], row_map["PC-006"], row_map["PC-007"], row_map["PC-008"], row_map["PC-009"], row_map["PC-010"]]
    stru11 = [row_map["PC-011"], row_map["PC-012"], row_map["PC-013"], row_map["PC-014"], row_map["PC-015"], row_map["PC-016"], row_map["PC-017"]]
    testing_cam = [row_map["PC-018"], row_map["PC-019"], row_map["PC-020"], row_map["PC-022"], row_map["PC-023"], row_map["PC-024"], row_map["PC-025"], row_map["PC-026"], row_map["PC-027"]]
    workflow_items = load_workflow_items()
    workflow_counts = Counter(classify_workflow_item(item)[0] for item in workflow_items)

    client_confirmation = [
        "Whether the STRU 11 AOS 2 bypass scenario and the 100-hour annual cap reflect actual maintenance practice and the intended flexibility the client wants preserved in the final permit.",
        "Whether the client can comply with the new effective-fence-line, patrol, camera, fence inspection, and breach-documentation obligations without additional staffing, contractor support, or capital work.",
        "Whether the reduced STRU 11 DDGS throughput limit of 178.8 tph and the higher 145 gpm recirculating scrubber-water CAM threshold are operationally workable under current plant practice.",
        "Whether MPCA intentionally removed the issued permit's PM10 and NOx/PM2.5 EBD modeling trigger/protocol frameworks and the changed STRU 14 periodic SO2 testing language, or whether those changes should be clarified or reinstated in permit comments.",
    ]

    def bullet_rows(selected_rows: list[RowSpec]) -> list[str]:
        out: list[str] = []
        for row in selected_rows:
            request_status, _, basis, priority = request_meta(row.comparison_id)
            out.append(
                f"- `{row.comparison_id}` ({row.new_section or row.old_section}): {row.initial_analysis} "
                f"Application basis: {request_status}; {basis} Recommended follow-up: {row.follow_up}."
            )
        return out

    lines = [
        "# Memorandum",
        "",
        "To:\tProject Team",
        "From:\tPermit Review Support",
        "Subject:\tDraft Permit Changes Summary",
        "Date:\tMarch 17, 2026",
        "",
        "This memorandum summarizes the pertinent permit changes identified by comparing the issued air permit "
        f"(`{OLD_PDF.name}`) to the draft major amendment permit (`{NEW_PDF.name}`). "
        "The goal is to facilitate internal review by focusing on substantive permit structure, condition language, monitoring, testing, and operational-limit changes rather than OCR noise or pagination shifts. This update also cross-checks the current application package to distinguish requested permit edits from additional agency changes that were not clearly requested. The accompanying workbook now includes a Validation Matrix and QA Spot Checks tabs so reviewers can see which conclusions are supported by direct condition-to-condition permit evidence and which items required broader structural comparison.",
        "",
        "This memorandum is organized into the following sections.",
        "",
        "Permit-wide Changes",
        "Total Facility / Modeling / Ambient Boundary",
        "STRU 11 DDGS Baghouse / Alternate Operating Scenario",
        "Testing, CAM, and Other Operating Parameter Updates",
        "Key Client Confirmation Items",
        "",
        "## Overall Conclusion",
        f"Overall outcome: {overall_outcome}",
        "",
        basis_paragraph,
        "",
        "## Workflow Cross-Check",
        "The uploaded standard AI workflow workbook was used as a completeness check against this consultant review. Most of its medium/high additional flags were not new substantive permit changes; they were overcalls caused by one-to-many condition restructuring, retained requirements being read as removed, or unchanged conditions being treated as additions. The main additional meaningful issue surfaced through that cross-check was the separate deletion of the issued permit's NOx/PM2.5 EBD modeling framework, which is now captured as `PC-028`.",
        "",
        f"Workflow cross-check counts: {', '.join(f'{k}: {v}' for k, v in sorted(workflow_counts.items()))}" if workflow_items else "Workflow cross-check workbook not available.",
        "",
        "## QA Reliability Note",
        f"{WORKFLOW_USE_NOTE}",
        "",
        "Manual spot-checking of COMG 2 confirmed the review comment is valid: the workflow workbook's `Removed & Added` and `Equipment Summaries` tabs overstated removals. Direct comparison of issued and draft requirements 5.2.9, 5.2.10, and 5.2.11 shows those conditions remain in the draft permit at the same requirement numbers, with only formatting/OCR-level differences. That means the consultant workbook should be treated as the primary review set, while the workflow workbook should be treated as a broad completeness screen only.",
        "",
        "## Permit-wide Changes",
        "The draft makes several broad changes that affect the overall permit structure and review posture before getting to individual requirement-level revisions.",
    ]
    lines.extend(bullet_rows(permit_wide))
    lines.extend(
        [
            "",
            "## Total Facility / Modeling / Ambient Boundary",
            "MPCA appears to have replaced the issued permit's pollutant-specific modeling trigger/protocol frameworks for PM10, NOx, and PM2.5 with a broader ambient-boundary and public-access control framework. That shift is one of the most important permit-wide changes because it affects how future modifications may be evaluated and creates new site-control and reporting obligations.",
        ]
    )
    lines.extend(bullet_rows(tfac_modeling))
    lines.extend(
        [
            "",
            "## STRU 11 DDGS Baghouse / Alternate Operating Scenario",
            "The draft's most consequential operational change is the addition of an alternate operating scenario for STRU 11 that would allow limited direct venting during TREA 2 maintenance. The related emission-limit, hour-cap, throughput, and testing provisions should be treated as a connected package during client review.",
        ]
    )
    lines.extend(bullet_rows(stru11))
    lines.extend(
        [
            "",
            "## Testing, CAM, and Other Operating Parameter Updates",
            "In addition to the STRU 11 revisions, the draft updates several testing schedules, testing methods, throughput values, and CAM indicator thresholds. Some of these appear manageable and test-supported, but they still change the enforceable permit basis and should be confirmed against actual operation and test history.",
        ]
    )
    lines.extend(bullet_rows(testing_cam))
    lines.extend(
        [
            "",
            "## Top Review Items",
        ]
    )
    lines.extend(
        f"- `{row.comparison_id}` ({row.new_section or row.old_section}) [{request_meta(row.comparison_id)[3]}]: {row.initial_analysis}"
        for row in top_rows
    )
    lines.extend(
        [
            "",
            "## Key Client Confirmation Items",
        ]
    )
    lines.extend(f"- {item}" for item in client_confirmation)
    lines.extend(
        [
            "",
            "## Items That Appear Editorial or Structural Only",
        ]
    )
    lines.extend(
        f"- `{row.comparison_id}` ({row.new_section or row.old_section}): {row.initial_analysis}"
        for row in editorial_rows[:8]
    )

    MD_OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = build_rows()
    significant_count = sum(1 for row in rows if row.conclusion == "Significant changes")
    overall_outcome = "Significant changes"
    basis_paragraph = (
        "The draft contains multiple material revisions that go beyond editorial cleanup. Review of all four application packages indicates Valero principally requested the STRU 11/TREA 2 maintenance-routing change and associated 100-hour limit, while the two DDGS drag minor amendments expressly said STRU 11 throughput was not being changed and the TO inspection minor amendment was limited to TREA 16 inspection language. Against that application backdrop, the most important unrequested or insufficiently supported draft changes are the replacement of the issued permit's PM10 and NOx/PM2.5 modeling trigger/protocol frameworks with a general public preclusion and boundary-control regime, the new ambient-boundary monitoring and reporting obligations, the tighter STRU 11 throughput limit, revised STRU 13/TREA 4 test- and CAM-based thresholds, and the changed STRU 14 periodic SO2 testing language. Manual QA also confirmed that the workflow workbook can generate false positives, so the consultant comparison and QA tabs should control over the workflow summary tabs."
    )
    workbook_path = write_workbook(rows, overall_outcome, basis_paragraph)
    write_memo(rows, overall_outcome, basis_paragraph)

    print(f"Saved spreadsheet: {workbook_path}")
    print(f"Saved narrative memo: {MD_OUT}")
    print(f"Comparison rows created: {len(rows)}")
    print(f"Overall outcome: {overall_outcome}")
    print(f"Rows classified as Significant changes: {significant_count}")


if __name__ == "__main__":
    main()
