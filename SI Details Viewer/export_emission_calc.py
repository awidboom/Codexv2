import argparse
import os
from collections import OrderedDict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from shared_config import BASE_DIR, load_config, resolve_baseline_workbook

def normalize_header(value):
    return " ".join(str(value or "").strip().lower().split())

def normalize_equi(value):
    return str(value or "").replace(" ", "").upper()

def to_float(value):
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None

def load_references(data_file):
    ref = pd.read_excel(data_file, sheet_name="References - PTE", header=None, engine="openpyxl")
    mapping = {}
    for _, row in ref.iterrows():
        first = row.iloc[0]
        second = row.iloc[1]
        if isinstance(first, (int, float)) and not pd.isna(first) and isinstance(second, str) and second.strip():
            mapping[str(int(first))] = second.strip()
    return mapping

def load_segments(data_file):
    df = pd.read_excel(data_file, sheet_name="PTE - Segment Emissions Calcs", header=5, engine="openpyxl")
    columns = {normalize_header(col): col for col in df.columns}
    def get_col(name):
        return columns.get(normalize_header(name))
    rows = []
    for _, row in df.iterrows():
        equi = row.get(get_col("TEMPO EQUI #"))
        if pd.isna(equi):
            equi = row.get(get_col("TEMPO EQUI # Lookup"))
        if pd.isna(equi):
            equi = row.get(get_col("TEMPO SEG VLOOKUP"))
        if pd.isna(equi):
            continue
        equi_str = str(equi).strip()
        if equi_str:
            equi_str = equi_str.split("-")[0].strip()
        rows.append({
            "equiId": equi_str,
            "seg": row.get(get_col("Seg")),
            "segDescription": row.get(get_col("Seg Description")),
            "sourceCategory": row.get(get_col("Source Category")),
            "pollutant": row.get(get_col("Pollutant")),
            "emissionFactor": row.get(get_col("Emission Factor (lb/unit)")),
            "emissionFactorUnit": row.get(get_col("Emission Factor Unit")),
            "emissionFactorSource": row.get(get_col("Emission Factor Source")),
            "efReference": row.get(get_col("EF Reference")),
            "throughput": row.get(get_col("Throughput (unit/hr)")),
            "throughputUnits": row.get(get_col("Throughput Units")),
            "hourlyPotential": row.get(get_col("Hourly Potential Emissions (lb/hr)")),
            "limitedPotential": row.get(get_col("Limited Potential Emissions (tpy)")),
            "tempoStru": row.get(get_col("TEMPO STRU #")),
            "tempoTrea": row.get(get_col("TEMPO TREA #")),
            "deltaSv": row.get(get_col("DELTA SV #")),
        })
    return rows

def choose_reference_id(row):
    source = row.get("emissionFactorSource")
    ef_ref = row.get("efReference")
    if isinstance(source, (int, float)) and not pd.isna(source):
        return str(int(source))
    if isinstance(ef_ref, (int, float)) and not pd.isna(ef_ref):
        return str(int(ef_ref))
    return None

def build_inputs(rows):
    throughput_map = OrderedDict()
    ef_map = OrderedDict()
    for row in rows:
        throughput = row.get("throughput")
        throughput_units = row.get("throughputUnits")
        if throughput is not None and not pd.isna(throughput) and throughput_units:
            throughput_map.setdefault((str(throughput).strip(), str(throughput_units).strip()), None)
        ef_value = row.get("emissionFactor")
        ef_unit = row.get("emissionFactorUnit")
        pollutant = row.get("pollutant")
        if ef_value is not None and not pd.isna(ef_value) and ef_unit:
            ef_map.setdefault((str(pollutant).strip(), str(ef_value).strip(), str(ef_unit).strip()), None)
    return throughput_map, ef_map

def apply_table_borders(ws, min_row, max_row, min_col, max_col):
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border

def write_workbook(rows, references, output_path, client, project, title):
    wb = Workbook()
    ws = wb.active
    ws.title = "Emission Calculations"
    for ref, text in zip(("A1", "A2", "A3"), (client, project, title)):
        ws[ref] = text
        ws[ref].font = Font(bold=True)
    input_row = 5
    ws[f"A{input_row}"] = "Inputs"
    ws[f"A{input_row}"].font = Font(bold=True)
    input_row += 1
    hours_cell = f"B{input_row}"
    ws[f"A{input_row}"] = "Hours of Operation"
    ws[hours_cell] = 8760
    ws[f"C{input_row}"] = "hr/yr"
    input_row += 1
    throughput_map, ef_map = build_inputs(rows)
    for (value, unit) in throughput_map:
        ws[f"A{input_row}"] = f"Throughput ({unit})"
        ws[f"B{input_row}"] = to_float(value)
        ws[f"C{input_row}"] = unit
        throughput_map[(value, unit)] = f"B{input_row}"
        input_row += 1
    for (pollutant, value, unit) in ef_map:
        ws[f"A{input_row}"] = f"EF {pollutant} ({unit})"
        ws[f"B{input_row}"] = to_float(value)
        ws[f"C{input_row}"] = unit
        ef_map[(pollutant, value, unit)] = f"B{input_row}"
        input_row += 1
    headers = ["Seg Description", "Source Category", "Pollutant", "Emission Factor", "EF Unit", "Throughput", "Throughput Units", "Emission Rate (lb/hr)", "Limited Potential (tpy)", "EF Source", "Reference #", "SV/STRU/TREA", "Review"]
    start_row = input_row + 2
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    footnotes = OrderedDict()
    mw_lookup = {"co": 28, "nox": 46, "no2": 46}
    row_idx = start_row + 1
    for row in rows:
        pollutant = str(row.get("pollutant") or "").strip()
        ef_value = row.get("emissionFactor")
        ef_unit = str(row.get("emissionFactorUnit") or "").strip()
        throughput_value = row.get("throughput")
        throughput_unit = str(row.get("throughputUnits") or "").strip()
        ws.cell(row=row_idx, column=1, value=row.get("segDescription"))
        ws.cell(row=row_idx, column=2, value=row.get("sourceCategory"))
        ws.cell(row=row_idx, column=3, value=pollutant)
        ws.cell(row=row_idx, column=5, value=ef_unit)
        ws.cell(row=row_idx, column=7, value=throughput_unit)
        ef_key = (pollutant, str(ef_value).strip(), ef_unit)
        ef_cell = ef_map.get(ef_key)
        ws.cell(row=row_idx, column=4, value=f"={ef_cell}" if ef_cell else to_float(ef_value))
        throughput_key = (str(throughput_value).strip(), throughput_unit)
        throughput_cell = throughput_map.get(throughput_key)
        ws.cell(row=row_idx, column=6, value=f"={throughput_cell}" if throughput_cell else to_float(throughput_value))
        review_flag = ""
        emission_rate_cell = ws.cell(row=row_idx, column=8)
        limited_cell = ws.cell(row=row_idx, column=9)
        ef_unit_l = ef_unit.lower()
        throughput_unit_l = throughput_unit.lower()
        if ef_unit_l == "ppm" and throughput_unit_l == "acfm":
            mw = mw_lookup.get(pollutant.lower())
            if mw:
                emission_rate_cell.value = f"=D{row_idx}*1E-6*(F{row_idx}*60/385.3)*{mw}"
                limited_cell.value = f"=H{row_idx}*{hours_cell}/2000"
            else:
                emission_rate_cell.value = to_float(row.get("hourlyPotential"))
                limited_cell.value = to_float(row.get("limitedPotential"))
                review_flag = "CHECK MW"
        elif "lb/mmbtu" in ef_unit_l and "mmbtu/hr" in throughput_unit_l:
            emission_rate_cell.value = f"=D{row_idx}*F{row_idx}"
            limited_cell.value = f"=H{row_idx}*{hours_cell}/2000"
        elif "lb/1000 lb" in ef_unit_l and "1000 lb" in throughput_unit_l:
            emission_rate_cell.value = f"=D{row_idx}*F{row_idx}"
            limited_cell.value = f"=H{row_idx}*{hours_cell}/2000"
        else:
            emission_rate_cell.value = to_float(row.get("hourlyPotential"))
            limited_cell.value = to_float(row.get("limitedPotential"))
            review_flag = "CHECK METHOD"
        if review_flag:
            emission_rate_cell.fill = review_fill
            limited_cell.fill = review_fill
            ws.cell(row=row_idx, column=13, value=review_flag)
        ws.cell(row=row_idx, column=10, value=row.get("emissionFactorSource"))
        ref_id = choose_reference_id(row)
        ref_text = references.get(ref_id) if ref_id else None
        if ref_text:
            footnotes.setdefault(ref_text, len(footnotes) + 1)
            ws.cell(row=row_idx, column=11, value=footnotes[ref_text])
        links = [row.get("deltaSv"), row.get("tempoStru"), row.get("tempoTrea")]
        ws.cell(row=row_idx, column=12, value=" / ".join([str(v) for v in links if v and not pd.isna(v)]))
        row_idx += 1
    end_row = row_idx - 1
    apply_table_borders(ws, start_row, end_row, 1, len(headers))
    widths = {"A": 24, "B": 18, "C": 14, "D": 16, "E": 12, "F": 14, "G": 16, "H": 18, "I": 18, "J": 16, "K": 10, "L": 18, "M": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = f"A{start_row + 1}"
    footnote_row = end_row + 2
    ws.cell(row=footnote_row, column=1, value="Footnotes").font = Font(bold=True)
    footnote_row += 1
    for text, number in footnotes.items():
        ws.cell(row=footnote_row, column=1, value=f"{number} - {text}")
        footnote_row += 1
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

def main():
    config = load_config()
    parser = argparse.ArgumentParser()
    parser.add_argument("--equi", required=True)
    parser.add_argument("--client", default=config.get("client_name", "Client Name"))
    parser.add_argument("--project", default=config.get("project_name", "Project Name"))
    parser.add_argument("--title", default=None)
    parser.add_argument("--output", default=None)
    args = parser.parse_args()
    data_file = resolve_baseline_workbook(config)
    segments = load_segments(data_file)
    references = load_references(data_file)
    rows = [row for row in segments if normalize_equi(row["equiId"]) == normalize_equi(args.equi)]
    if not rows:
        raise SystemExit(f"No segments found for {args.equi}.")
    output = args.output or os.path.join(BASE_DIR, "exports", f"{normalize_equi(args.equi)}_emission_calcs.xlsx")
    title = args.title or f"Emission Calculations for {args.equi}"
    write_workbook(rows, references, output, args.client, args.project, title)
    print(f"Wrote {output}")

if __name__ == "__main__":
    main()
