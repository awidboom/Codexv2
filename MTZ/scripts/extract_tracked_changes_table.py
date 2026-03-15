import argparse
import datetime as dt
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NS = {"w": W_NS}


def _local(tag: str) -> str:
    return tag.split("}", 1)[1] if "}" in tag else tag


def _w_attr(name: str) -> str:
    return f"{{{W_NS}}}{name}"


def _normalize_space(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def _text_from_container(node: ET.Element, *, include_deleted: bool) -> str:
    parts: list[str] = []
    for el in node.iter():
        t = _local(el.tag)
        if t == "t" and el.text:
            parts.append(el.text)
        elif include_deleted and t == "delText" and el.text:
            parts.append(el.text)
        elif t == "tab":
            parts.append("\t")
        elif t == "br":
            parts.append("\n")
    return "".join(parts)


def _paragraph_style_id(p: ET.Element) -> str | None:
    p_style = p.find("w:pPr/w:pStyle", NS)
    if p_style is None:
        return None
    return p_style.get(_w_attr("val"))


def _is_heading_style(style_id: str | None) -> bool:
    return bool(style_id) and style_id.lower().startswith("heading")


_RE_CONDITION = re.compile(r"\b(?:Permit\s+)?Condition\s+(\d{3,})\b", re.I)
_RE_TABLE_CAPTION_START = re.compile(
    r"^\s*Table\s+([IVXLCDM]+\.[A-Za-z0-9.\-]+|[A-Za-z0-9][\w.\-]*)\b", re.I
)
_RE_PART = re.compile(r"\bPart\s+(\d+)\b", re.I)
_RE_SECTION = re.compile(r"^\s*Section\s+([IVXLCDM]+)\b", re.I)
_RE_TABLE_CODE = re.compile(r"\bTable\s+([IVXLCDM]+)\s*[–-]\s*([A-Z]\.\d+(?:\.\d+)*)\b")
_RE_TABLE_SIMPLE_CODE = re.compile(r"\bTable\s+([IVXLCDM]+)\s+([A-Z0-9][A-Za-z0-9.\-]*)\b")


def _section_from_heading(text: str, style_id: str) -> str:
    if (style_id or "") != "Heading1":
        return ""
    t = _normalize_space(text).upper()
    if t == "PERMIT CONDITIONS":
        return "VI"
    if "APPLICABLE LIMITS" in t and "COMPLIANCE MONITORING" in t:
        return "VII"
    return ""


def _extract_consent_decree_citations(text: str) -> list[str]:
    patterns = [
        r"\(Basis:\s*[^)]*Consent Decree[^)]*\)",
        r"\(basis:\s*[^)]*Consent Decree[^)]*\)",
        r"\bBasis:\s*[^.\n]{0,250}Consent Decree[^.\n]{0,250}[.)]",
        r"\bConsent Decree\s+paragraph\s+\d+\b",
        r"\bConsent Decree\s*\(?Tesoro\)?\s*Appendix\s+[A-Za-z0-9\-]+,\s*Condition\s+\d+\b",
        r"\bConsent Decree\s+\d{3,}\b",
    ]
    out: list[str] = []
    for pat in patterns:
        for m in re.finditer(pat, text, re.I):
            s = _normalize_space(m.group(0))
            if s and s not in out:
                out.append(s)
    return out


def _best_citation(window_texts: list[tuple[int, str]], *, center_idx: int) -> str:
    candidates: list[tuple[int, int, str]] = []
    for para_idx, t in window_texts:
        for c in _extract_consent_decree_citations(t):
            score = 0
            cl = c.lower()
            if "paragraph" in cl:
                score += 10
            if "appendix" in cl:
                score += 7
            if "condition" in cl:
                score += 3
            dist = abs(para_idx - center_idx)
            candidates.append((score, -dist, c))
    if not candidates:
        return ""
    candidates.sort(reverse=True)
    return candidates[0][2]


@dataclass(frozen=True)
class Anchors:
    heading: str = ""
    condition: str = ""
    table_caption: str = ""
    condition_last_citation: str = ""


@dataclass(frozen=True)
class LvlDef:
    ilvl: int
    num_fmt: str
    lvl_text: str
    start: int


def _to_roman(n: int) -> str:
    vals = [
        (1000, "M"),
        (900, "CM"),
        (500, "D"),
        (400, "CD"),
        (100, "C"),
        (90, "XC"),
        (50, "L"),
        (40, "XL"),
        (10, "X"),
        (9, "IX"),
        (5, "V"),
        (4, "IV"),
        (1, "I"),
    ]
    out: list[str] = []
    for v, s in vals:
        while n >= v:
            out.append(s)
            n -= v
    return "".join(out) if out else "I"


def _to_alpha(n: int, *, upper: bool) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(ord("A") + r) + s
    return s if upper else s.lower()


def _format_num(n: int, num_fmt: str) -> str:
    if num_fmt in ("decimal", "decimalZero"):
        return str(n)
    if num_fmt == "upperRoman":
        return _to_roman(n)
    if num_fmt == "lowerRoman":
        return _to_roman(n).lower()
    if num_fmt == "upperLetter":
        return _to_alpha(n, upper=True)
    if num_fmt == "lowerLetter":
        return _to_alpha(n, upper=False)
    if num_fmt == "bullet":
        return "•"
    return str(n)


class NumberingResolver:
    def __init__(self, numbering_root: ET.Element | None):
        self._num_to_abstract: dict[str, str] = {}
        self._abstract_lvls: dict[str, dict[int, LvlDef]] = {}
        self._start_override: dict[tuple[str, int], int] = {}
        self._counters: dict[str, list[int]] = {}

        if numbering_root is None:
            return

        for abs_num in numbering_root.findall("w:abstractNum", NS):
            abs_id = abs_num.get(_w_attr("abstractNumId"))
            if abs_id is None:
                continue
            lvls: dict[int, LvlDef] = {}
            for lvl in abs_num.findall("w:lvl", NS):
                ilvl_s = lvl.get(_w_attr("ilvl"))
                if ilvl_s is None:
                    continue
                ilvl = int(ilvl_s)
                num_fmt_el = lvl.find("w:numFmt", NS)
                lvl_text_el = lvl.find("w:lvlText", NS)
                start_el = lvl.find("w:start", NS)
                num_fmt = num_fmt_el.get(_w_attr("val")) if num_fmt_el is not None else "decimal"
                lvl_text = lvl_text_el.get(_w_attr("val")) if lvl_text_el is not None else ""
                start = int(start_el.get(_w_attr("val"))) if start_el is not None and start_el.get(_w_attr("val")) else 1
                lvls[ilvl] = LvlDef(ilvl=ilvl, num_fmt=num_fmt or "decimal", lvl_text=lvl_text or "", start=start)
            self._abstract_lvls[abs_id] = lvls

        for num in numbering_root.findall("w:num", NS):
            num_id = num.get(_w_attr("numId"))
            abs_id_el = num.find("w:abstractNumId", NS)
            abs_id = abs_id_el.get(_w_attr("val")) if abs_id_el is not None else None
            if num_id and abs_id:
                self._num_to_abstract[num_id] = abs_id

            for lvl_ovr in num.findall("w:lvlOverride", NS):
                ilvl_s = lvl_ovr.get(_w_attr("ilvl"))
                if ilvl_s is None or num_id is None:
                    continue
                start_ovr = lvl_ovr.find("w:startOverride", NS)
                if start_ovr is None:
                    continue
                v = start_ovr.get(_w_attr("val"))
                if v:
                    self._start_override[(num_id, int(ilvl_s))] = int(v)

    def reset(self) -> None:
        self._counters.clear()

    def label_for_paragraph(self, p: ET.Element) -> str:
        num_pr = p.find("w:pPr/w:numPr", NS)
        if num_pr is None:
            return ""
        num_id_el = num_pr.find("w:numId", NS)
        ilvl_el = num_pr.find("w:ilvl", NS)
        if num_id_el is None or ilvl_el is None:
            return ""
        num_id = num_id_el.get(_w_attr("val"))
        ilvl_s = ilvl_el.get(_w_attr("val"))
        if not num_id or ilvl_s is None:
            return ""
        ilvl = int(ilvl_s)

        abs_id = self._num_to_abstract.get(num_id)
        lvl_defs = self._abstract_lvls.get(abs_id or "", {})
        lvl_def = lvl_defs.get(ilvl)
        if lvl_def is None:
            return ""

        if num_id not in self._counters:
            self._counters[num_id] = [0] * 9

        start = self._start_override.get((num_id, ilvl), lvl_def.start)
        if self._counters[num_id][ilvl] == 0:
            self._counters[num_id][ilvl] = start
        else:
            self._counters[num_id][ilvl] += 1
        for deeper in range(ilvl + 1, 9):
            self._counters[num_id][deeper] = 0

        tmpl = lvl_def.lvl_text
        if not tmpl:
            return _format_num(self._counters[num_id][ilvl], lvl_def.num_fmt) + "."

        def repl(m: re.Match) -> str:
            level = int(m.group(1)) - 1
            if level < 0 or level >= 9:
                return m.group(0)
            value = self._counters[num_id][level]
            if value == 0:
                value = start if level == ilvl else 1
            fmt = lvl_defs.get(level).num_fmt if lvl_defs.get(level) else "decimal"
            return _format_num(value, fmt)

        rendered = re.sub(r"%([1-9])", repl, tmpl)
        rendered = rendered.replace("\uF0B7", "•")
        return rendered.strip()


def _update_anchors_for_paragraph(
    anchors: Anchors,
    p: ET.Element,
    *,
    in_table: bool,
    numbering: NumberingResolver,
    reset_numbering_on_condition: bool,
) -> tuple[Anchors, str, str, str, str]:
    style_id = _paragraph_style_id(p)
    text = _normalize_space(_text_from_container(p, include_deleted=False))

    heading = anchors.heading
    if _is_heading_style(style_id) and text:
        heading = f"Heading: {text[:120]}"

    condition = anchors.condition
    condition_changed = False
    # Only treat this paragraph as setting the active Condition when it looks like a true Condition header,
    # not a cross-reference (e.g., "Refer to Condition 24324...").
    looks_like_condition_header = False
    if (style_id or "") == "Heading2":
        looks_like_condition_header = bool(re.match(r"^\s*Condition\s+\d{3,}\b", text, re.I))
    else:
        looks_like_condition_header = bool(re.match(r"^\s*(?:Permit\s+)?Condition\s+\d{3,}\b", text, re.I))
    if looks_like_condition_header:
        m = _RE_CONDITION.search(text)
        if m:
            condition = f"Condition {m.group(1)}"
            condition_changed = True

    if reset_numbering_on_condition and condition_changed:
        numbering.reset()

    table_caption = anchors.table_caption
    # Only treat "Table ..." as a *caption* when it's outside the table (or explicitly styled as a caption),
    # to avoid accidentally capturing regulatory text like "table 3 to this subpart..." inside table cells.
    if not in_table:
        m2 = _RE_TABLE_CAPTION_START.search(text)
        style_l = (style_id or "").lower() if style_id else ""
        looks_like_caption = bool(m2) or ("caption" in style_l and "table" in text.lower())
        if looks_like_caption and "table" in text.lower():
            table_caption = text[:200]

    list_label = numbering.label_for_paragraph(p)
    part_label = ""
    mpart = _RE_PART.search(text)
    if mpart:
        part_label = f"Part {mpart.group(1)}"

    condition_last_citation = anchors.condition_last_citation
    cites = _extract_consent_decree_citations(text)
    if cites:
        condition_last_citation = _best_citation([(0, text)], center_idx=0)

    return (
        Anchors(
            heading=heading,
            condition=condition,
            table_caption=table_caption,
            condition_last_citation=condition_last_citation,
        ),
        text,
        style_id or "",
        part_label,
        list_label,
    )


def _extract_changes(p: ET.Element) -> list[dict]:
    out: list[dict] = []
    for el in p.iter():
        t = _local(el.tag)
        if t not in ("ins", "del"):
            continue
        if t == "ins":
            change_text = _normalize_space(_text_from_container(el, include_deleted=False))
            change_type = "INSERT"
        else:
            change_text = _normalize_space(_text_from_container(el, include_deleted=True))
            change_type = "DELETE"
        if not change_text:
            continue
        out.append({"change_type": change_type, "change_text": change_text})
    return out


def _extract_revision_texts(p: ET.Element) -> tuple[str, str]:
    ins_parts: list[str] = []
    del_parts: list[str] = []
    for el in p.iter():
        t = _local(el.tag)
        if t == "ins":
            ins_parts.append(_text_from_container(el, include_deleted=False))
        elif t == "del":
            del_parts.append(_text_from_container(el, include_deleted=True))
    return _normalize_space(" ".join(ins_parts)), _normalize_space(" ".join(del_parts))


def _try_load_numbering_root(docx_path: str) -> ET.Element | None:
    try:
        with zipfile.ZipFile(docx_path) as z:
            return ET.fromstring(z.read("word/numbering.xml"))
    except KeyError:
        return None


def _extract_table_cells(tbl: ET.Element) -> list[list[ET.Element]]:
    rows: list[list[ET.Element]] = []
    for tr in tbl.findall("w:tr", NS):
        rows.append(tr.findall("w:tc", NS))
    return rows


def _cell_text(tc: ET.Element) -> str:
    return _normalize_space(_text_from_container(tc, include_deleted=False))


def _infer_table_caption_from_header(header_cells: list[ET.Element]) -> str:
    # Many permit tables store the visible "Table X – ..." label in the header row,
    # not in a separate caption paragraph. Prefer that when present.
    for tc in header_cells:
        t = _cell_text(tc)
        if not t:
            continue
        m = re.search(r"\bTable\b", t, re.I)
        if not m:
            continue
        cap = _normalize_space(t[m.start() :])[:200]
        if cap.lower().startswith("table"):
            return cap
    return ""


def _make_location(
    *,
    condition: str,
    part_label: str,
    list_label: str,
    in_table: bool,
    table_caption: str,
    table_idx: str,
    table_row_hint: str,
    table_col_hint: str,
    heading: str,
    para_idx: int,
) -> str:
    parts: list[str] = []
    if condition:
        parts.append(condition)
    if part_label:
        parts.append(part_label)
    if list_label:
        parts.append(list_label)
    if in_table:
        if table_caption:
            parts.append(table_caption)
        elif table_idx:
            parts.append(f"Table (unlabeled #{table_idx})")
        if table_row_hint:
            parts.append(f"Row: {table_row_hint}")
        if table_col_hint:
            parts.append(f"Col: {table_col_hint}")
    if not parts:
        if heading:
            parts.append(heading)
        else:
            parts.append(f"Paragraph {para_idx}")
    return " - ".join(parts)


def build_rows(
    docx_path: str,
    *,
    window: int,
    reset_numbering_on_condition: bool,
) -> tuple[list[dict], list[dict]]:
    with zipfile.ZipFile(docx_path) as z:
        root = ET.fromstring(z.read("word/document.xml"))

    body = root.find("w:body", NS)
    if body is None:
        raise RuntimeError("Could not find word/document.xml w:body.")

    numbering = NumberingResolver(_try_load_numbering_root(docx_path))

    paragraphs: list[dict] = []
    changes: list[dict] = []
    anchors = Anchors()

    para_idx = 0
    table_idx = 0
    global_last_citation = ""
    current_section = ""

    for child in list(body):
        kind = _local(child.tag)
        if kind == "p":
            in_table = False
            anchors, text, style_id, part_label, list_label = _update_anchors_for_paragraph(
                anchors,
                child,
                in_table=in_table,
                numbering=numbering,
                reset_numbering_on_condition=reset_numbering_on_condition,
            )
            if anchors.condition_last_citation:
                global_last_citation = anchors.condition_last_citation

            sec = _section_from_heading(text, style_id)
            if sec:
                current_section = sec

            paragraphs.append(
                {
                    "para_idx": para_idx,
                    "text": text,
                    "style_id": style_id,
                    "heading": anchors.heading,
                    "condition": anchors.condition,
                    "part_label": part_label,
                    "list_label": list_label,
                    "section": current_section,
                    "in_table": False,
                    "table_idx": "",
                    "table_caption": anchors.table_caption,
                    "table_row_hint": "",
                    "table_col_hint": "",
                    "condition_last_citation": anchors.condition_last_citation,
                }
            )
            for c in _extract_changes(child):
                changes.append({"para_idx": para_idx, "in_table": False, "table_idx": "", **c})
            para_idx += 1
            continue

        if kind == "tbl":
            table_idx += 1
            grid = _extract_table_cells(child)
            header_cells = grid[0] if grid else []
            table_caption_for_table = _infer_table_caption_from_header(header_cells) or anchors.table_caption
            col_hints = [_cell_text(tc)[:80] for tc in header_cells]

            for row in grid:
                row_hint = _cell_text(row[0])[:120] if row else ""
                for c_i, tc in enumerate(row):
                    col_hint = col_hints[c_i] if c_i < len(col_hints) else ""
                    for p in tc.findall(".//w:p", NS):
                        in_table = True
                        anchors, text, style_id, part_label, list_label = _update_anchors_for_paragraph(
                            anchors,
                            p,
                            in_table=in_table,
                            numbering=numbering,
                            reset_numbering_on_condition=reset_numbering_on_condition,
                        )
                        if anchors.condition_last_citation:
                            global_last_citation = anchors.condition_last_citation

                        paragraphs.append(
                            {
                                "para_idx": para_idx,
                                "text": text,
                                "style_id": style_id,
                                "heading": anchors.heading,
                                "condition": anchors.condition,
                                "part_label": part_label,
                                "list_label": list_label,
                                "section": current_section,
                                "in_table": True,
                                "table_idx": str(table_idx),
                                "table_caption": table_caption_for_table,
                                "table_row_hint": row_hint,
                                "table_col_hint": col_hint,
                                "condition_last_citation": anchors.condition_last_citation,
                            }
                        )
                        for c in _extract_changes(p):
                            changes.append({"para_idx": para_idx, "in_table": True, "table_idx": str(table_idx), **c})
                        para_idx += 1

    para_by_idx = {p["para_idx"]: p for p in paragraphs}

    out_rows: list[dict] = []
    for i, ch in enumerate(changes, 1):
        center = ch["para_idx"]
        p = para_by_idx.get(center)
        if not p:
            continue

        window_texts: list[tuple[int, str]] = []
        for j in range(max(0, center - window), center + window + 1):
            pj = para_by_idx.get(j)
            if pj and pj["text"]:
                window_texts.append((j, pj["text"]))
        citation = _best_citation(window_texts, center_idx=center)
        if not citation:
            citation = p.get("condition_last_citation", "") or global_last_citation

        location = _make_location(
            condition=p.get("condition", ""),
            part_label=p.get("part_label", ""),
            list_label=p.get("list_label", ""),
            in_table=bool(ch.get("in_table")),
            table_caption=p.get("table_caption", ""),
            table_idx=p.get("table_idx", ""),
            table_row_hint=p.get("table_row_hint", ""),
            table_col_hint=p.get("table_col_hint", ""),
            heading=p.get("heading", ""),
            para_idx=center,
        )

        out_rows.append(
            {
                "row_id": i,
                "location_in_word": location,
                "consent_decree_citation": citation,
                "regulatory_language": f"{ch['change_type']}: {ch['change_text']}",
                "para_idx": center,
                "paragraph_excerpt": (p.get("text", "")[:240] if p.get("text") else ""),
            }
        )

    grouped: dict[tuple[str, str], dict] = {}
    for r in out_rows:
        key = (r["location_in_word"], r["consent_decree_citation"])
        if key not in grouped:
            grouped[key] = {
                "location_in_word": r["location_in_word"],
                "consent_decree_citation": r["consent_decree_citation"],
                "regulatory_language": [r["regulatory_language"]],
                "first_row_id": r["row_id"],
                "last_row_id": r["row_id"],
                "example_paragraph_excerpt": r["paragraph_excerpt"],
            }
        else:
            grouped[key]["regulatory_language"].append(r["regulatory_language"])
            grouped[key]["last_row_id"] = r["row_id"]

    grouped_rows: list[dict] = []
    for g in grouped.values():
        grouped_rows.append(
            {
                "location_in_word": g["location_in_word"],
                "consent_decree_citation": g["consent_decree_citation"],
                "regulatory_language": " | ".join(g["regulatory_language"]),
                "first_row_id": g["first_row_id"],
                "last_row_id": g["last_row_id"],
                "example_paragraph_excerpt": g["example_paragraph_excerpt"],
            }
        )

    def _sort_key(r: dict):
        loc = r.get("location_in_word", "")
        m = re.search(r"\bCondition\s+(\d{3,})\b", loc)
        if m:
            return (0, int(m.group(1)), loc)
        m2 = re.search(r"\bTable\s+([IVXLCDM]+\.[A-Za-z0-9.\-]+|[A-Za-z0-9][\w.\-]*)\b", loc, re.I)
        if m2:
            return (1, m2.group(1), loc)
        return (2, loc, loc)

    grouped_rows.sort(key=_sort_key)
    return out_rows, grouped_rows


def build_crosswalk(
    docx_path: str,
    *,
    window: int,
    reset_numbering_on_condition: bool,
) -> tuple[list[dict], list[dict]]:
    """
    Returns:
      - condition_paragraph_rows: Section VI only (Permit Conditions), paragraph-level crosswalk
      - table_rows: Other sections, table-level crosswalk (e.g., VII-C.2.1)
    """
    with zipfile.ZipFile(docx_path) as z:
        root = ET.fromstring(z.read("word/document.xml"))

    body = root.find("w:body", NS)
    if body is None:
        raise RuntimeError("Could not find word/document.xml w:body.")

    numbering = NumberingResolver(_try_load_numbering_root(docx_path))
    anchors = Anchors()
    current_section = ""
    global_last_citation = ""
    para_idx = 0
    table_idx = 0

    condition_paragraph_rows: list[dict] = []
    table_acc: dict[str, dict] = {}
    active_condition = ""
    active_para_label = ""

    def _table_key_from_caption(cap: str, table_idx: int) -> tuple[str, str]:
        cap_norm = _normalize_space(cap)
        m = _RE_TABLE_CODE.search(cap_norm)
        if m:
            return f"{m.group(1).upper()}-{m.group(2)}", m.group(1).upper()
        m2 = _RE_TABLE_SIMPLE_CODE.search(cap_norm)
        if m2:
            return f"{m2.group(1).upper()}-{m2.group(2)}", m2.group(1).upper()
        m3 = re.search(r"\bTable\s+([IVXLCDM]+)\b", cap_norm, re.I)
        if m3:
            roman = m3.group(1).upper()
            return roman, roman
        return f"Table#{table_idx}", ""

    def add_table_summary(
        table_key: str,
        *,
        cap: str,
        section_roman: str,
        change_count: int,
        involves_23129: bool,
        involves_23562: bool,
        involves_24324: bool,
    ) -> None:
        if table_key not in table_acc:
            table_acc[table_key] = {
                "table_key": table_key,
                "table_caption": cap,
                "section_roman": section_roman,
                "change_count": 0,
                "involves_23129": False,
                "involves_23562": False,
                "involves_24324": False,
            }
        table_acc[table_key]["table_caption"] = table_acc[table_key]["table_caption"] or cap
        table_acc[table_key]["section_roman"] = table_acc[table_key]["section_roman"] or section_roman
        table_acc[table_key]["change_count"] += change_count
        table_acc[table_key]["involves_23129"] = table_acc[table_key]["involves_23129"] or involves_23129
        table_acc[table_key]["involves_23562"] = table_acc[table_key]["involves_23562"] or involves_23562
        table_acc[table_key]["involves_24324"] = table_acc[table_key]["involves_24324"] or involves_24324

    for child in list(body):
        kind = _local(child.tag)
        if kind == "p":
            in_table = False
            anchors, text, style_id, part_label, list_label = _update_anchors_for_paragraph(
                anchors,
                child,
                in_table=in_table,
                numbering=numbering,
                reset_numbering_on_condition=reset_numbering_on_condition,
            )
            if anchors.condition_last_citation:
                global_last_citation = anchors.condition_last_citation

            sec = _section_from_heading(text, style_id)
            if sec:
                current_section = sec

            ins_text, del_text = _extract_revision_texts(child)
            cites_final = _extract_consent_decree_citations(text)
            cites_ins = _extract_consent_decree_citations(ins_text) if ins_text else []
            cites_del = _extract_consent_decree_citations(del_text) if del_text else []

            # Section VI: permit conditions, paragraph-level (only keep rows that touch Consent Decree).
            if current_section == "VI" and anchors.condition:
                if anchors.condition != active_condition:
                    active_condition = anchors.condition
                    active_para_label = ""
                if list_label:
                    active_para_label = list_label
                if cites_final or cites_ins or cites_del:
                    # Prefer citations local to the paragraph; if empty, fall back to nearby window/global within same condition.
                    citation = (
                        _best_citation([(para_idx, text)], center_idx=para_idx)
                        or anchors.condition_last_citation
                        or global_last_citation
                    )
                    condition_id = anchors.condition.replace("Condition ", "").strip()
                    location = _make_location(
                        condition=anchors.condition,
                        part_label=part_label,
                        list_label=active_para_label,
                        in_table=False,
                        table_caption="",
                        table_idx="",
                        table_row_hint="",
                        table_col_hint="",
                        heading=anchors.heading,
                        para_idx=para_idx,
                    )
                    condition_paragraph_rows.append(
                        {
                            "condition": condition_id,
                            "paragraph_label": active_para_label,
                            "location_in_word": location,
                            "consent_decree_current": "; ".join(cites_final) if cites_final else citation,
                            "consent_decree_added": "; ".join(cites_ins),
                            "consent_decree_removed": "; ".join(cites_del),
                        }
                    )

            para_idx += 1
            continue

        if kind == "tbl":
            table_idx += 1
            grid = _extract_table_cells(child)
            header_cells = grid[0] if grid else []
            # Only summarize tables for the non-VI sections the user cares about (esp. Section VII).
            cap = _infer_table_caption_from_header(header_cells) or anchors.table_caption
            table_key, table_section = _table_key_from_caption(cap, table_idx)
            # Filter to the sections requested (II, IV, VII) based on the table label itself.
            if table_section not in ("II", "IV", "VII"):
                for row in grid:
                    for tc in row:
                        para_idx += len(tc.findall(".//w:p", NS))
                continue

            change_count = 0
            involves_23129 = False
            involves_23562 = False
            involves_24324 = False

            for row in grid:
                for tc in row:
                    for p in tc.findall(".//w:p", NS):
                        anchors, text, style_id, part_label, list_label = _update_anchors_for_paragraph(
                            anchors,
                            p,
                            in_table=True,
                            numbering=numbering,
                            reset_numbering_on_condition=reset_numbering_on_condition,
                        )
                        if anchors.condition_last_citation:
                            global_last_citation = anchors.condition_last_citation

                        ins_text, del_text = _extract_revision_texts(p)
                        change_count += len(_extract_changes(p))
                        blob = " ".join([text, ins_text, del_text]).lower()
                        involves_23129 = involves_23129 or bool(re.search(r"\b23129\b", blob))
                        involves_23562 = involves_23562 or bool(re.search(r"\b23562\b", blob))
                        involves_24324 = involves_24324 or bool(re.search(r"\b24324\b", blob))
                        para_idx += 1
            add_table_summary(
                table_key,
                cap=_normalize_space(cap),
                section_roman=table_section,
                change_count=change_count,
                involves_23129=involves_23129,
                involves_23562=involves_23562,
                involves_24324=involves_24324,
            )
            continue

    # De-dupe Section VI rows: keep one per condition+paragraph_label.
    dedup: dict[tuple[str, str], dict] = {}
    for r in condition_paragraph_rows:
        key = (r["condition"], r["paragraph_label"])
        if key not in dedup:
            dedup[key] = r
        else:
            for f in ("consent_decree_added", "consent_decree_removed"):
                if r[f]:
                    dedup[key][f] = "; ".join([x for x in [dedup[key][f], r[f]] if x])
            if r["consent_decree_current"] and not dedup[key]["consent_decree_current"]:
                dedup[key]["consent_decree_current"] = r["consent_decree_current"]

    condition_paragraph_rows = list(dedup.values())
    condition_paragraph_rows.sort(key=lambda r: (int(r["condition"]), r["paragraph_label"] or ""))

    table_rows: list[dict] = []
    for v in table_acc.values():
        if v.get("change_count", 0) <= 0:
            continue
        table_rows.append(
            {
                "table_key": v["table_key"],
                "table_caption": v["table_caption"],
                "section_roman": v["section_roman"],
                "change_count": v["change_count"],
                "involves_23129": "Y" if v["involves_23129"] else "",
                "involves_23562": "Y" if v["involves_23562"] else "",
                "involves_24324": "Y" if v["involves_24324"] else "",
            }
        )
    table_rows.sort(key=lambda r: (r["section_roman"], r["table_key"]))

    return condition_paragraph_rows, table_rows


def _write_markdown(path: str, rows: list[dict], columns: list[str], *, max_rows: int) -> None:
    rows = rows[:max_rows]
    with open(path, "w", encoding="utf-8") as f:
        f.write("| " + " | ".join(columns) + " |\n")
        f.write("| " + " | ".join(["---"] * len(columns)) + " |\n")
        for r in rows:
            vals = []
            for c in columns:
                v = str(r.get(c, ""))
                v = v.replace("\n", " ").replace("|", "\\|")
                vals.append(v)
            f.write("| " + " | ".join(vals) + " |\n")


def _write_xlsx(path: str, per_change_rows: list[dict], grouped_rows: list[dict]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "per_change"
    ws2 = wb.create_sheet("grouped")

    per_cols = ["row_id", "location_in_word", "consent_decree_citation", "regulatory_language", "para_idx", "paragraph_excerpt"]
    grp_cols = [
        "location_in_word",
        "consent_decree_citation",
        "regulatory_language",
        "first_row_id",
        "last_row_id",
        "example_paragraph_excerpt",
    ]

    ws1.append(per_cols)
    for r in per_change_rows:
        ws1.append([r.get(c, "") for c in per_cols])

    ws2.append(grp_cols)
    for r in grouped_rows:
        ws2.append([r.get(c, "") for c in grp_cols])

    wb.save(path)


def _write_crosswalk_xlsx(path: str, condition_rows: list[dict], table_rows: list[dict]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Section_VI_Conditions"
    ws2 = wb.create_sheet("Other_Section_Tables")

    cond_cols = [
        "condition",
        "paragraph_label",
        "location_in_word",
        "consent_decree_current",
        "consent_decree_added",
        "consent_decree_removed",
    ]
    tab_cols = [
        "section_roman",
        "table_key",
        "table_caption",
        "change_count",
        "involves_23129",
        "involves_23562",
        "involves_24324",
    ]

    ws1.append(cond_cols)
    for r in condition_rows:
        ws1.append([r.get(c, "") for c in cond_cols])

    ws2.append(tab_cols)
    for r in table_rows:
        ws2.append([r.get(c, "") for c in tab_cols])

    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Extract tracked changes from a .docx and build an XLSX table with Word locations and Consent Decree citations."
    )
    ap.add_argument("docx", help="Input .docx with tracked changes")
    ap.add_argument("--out-dir", default="outputs", help="Output directory (default: outputs)")
    ap.add_argument(
        "--window",
        type=int,
        default=2,
        help="Paragraph window (+/-) to infer Consent Decree citations (default: 2)",
    )
    ap.add_argument(
        "--md-max-rows",
        type=int,
        default=50,
        help="Max rows to include in the markdown preview table (default: 50)",
    )
    ap.add_argument(
        "--reset-numbering-on-condition",
        action="store_true",
        help="Reset list numbering when a new 'Condition ####' paragraph is encountered.",
    )
    ap.add_argument(
        "--mode",
        choices=["changes", "crosswalk"],
        default="changes",
        help="Output mode: 'crosswalk' focuses on Section VI condition paragraphs and other-section table keys.",
    )
    args = ap.parse_args()

    os.makedirs(args.out_dir, exist_ok=True)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(os.path.basename(args.docx))[0]

    if args.mode == "changes":
        per_change_rows, grouped_rows = build_rows(
            args.docx,
            window=args.window,
            reset_numbering_on_condition=args.reset_numbering_on_condition,
        )
        out_xlsx = os.path.join(args.out_dir, f"{base}__tracked_changes__{ts}.xlsx")
        preview_md = os.path.join(args.out_dir, f"{base}__tracked_changes__preview__{ts}.md")
        _write_xlsx(out_xlsx, per_change_rows, grouped_rows)
        _write_markdown(
            preview_md,
            grouped_rows,
            columns=["location_in_word", "consent_decree_citation", "regulatory_language"],
            max_rows=args.md_max_rows,
        )
        print(out_xlsx)
        print(preview_md)
        print(f"Rows: {len(per_change_rows)}; Grouped: {len(grouped_rows)}")
        return 0

    condition_rows, table_rows = build_crosswalk(
        args.docx,
        window=args.window,
        reset_numbering_on_condition=args.reset_numbering_on_condition,
    )
    out_xlsx = os.path.join(args.out_dir, f"{base}__consent_decree_crosswalk__{ts}.xlsx")
    _write_crosswalk_xlsx(out_xlsx, condition_rows, table_rows)
    print(out_xlsx)
    print(f"Section VI rows: {len(condition_rows)}; Table rows: {len(table_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
