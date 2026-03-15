import argparse
import copy
import json
import re
from collections import defaultdict
from functools import lru_cache
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


ROOT_DIR = Path(__file__).resolve().parents[1]
SOURCE_MAP_PATH = ROOT_DIR / "tanks_main.js.map"

MONTH_LABELS = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
    "Annual",
]

MONTH_SHORT = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
    "Ann",
]

MONTH_TO_INDEX = {label.lower(): idx for idx, label in enumerate(MONTH_LABELS)}
MONTH_TO_INDEX.update({label.lower(): idx for idx, label in enumerate(MONTH_SHORT)})

TANK_ROW_KEYS = [
    "tankType",
    "tankIdentification",
    "location",
    "tankChar",
    "tankFit",
    "tankContents",
    "tanSolAbs",
    "petChem",
    "petDist",
]

FILLABLE_TANK_COLUMNS = [
    "tankType.tanTyp",
    "tankIdentification.tankID",
    "tankIdentification.tankDescription",
    "tankIdentification.tankCity",
    "tankIdentification.tankState",
    "tankIdentification.company",
    "location.loc",
    "tankChar.sheLen",
    "tankChar.sheHei",
    "tankChar.sheDia",
    "tankChar.maxLiqHei",
    "tankChar.avgLiqHei",
    "tankChar.minLiqHei",
    "tankChar.tanHea",
    "tankChar.maxHeaTem",
    "tankChar.avgHeaTem",
    "tankChar.minHeaTem",
    "tankChar.heaCyc",
    "tankChar.rooTyp",
    "tankChar.vacSet",
    "tankChar.preSet",
    "tankChar.vapSpaPre",
    "tankChar.tanIns",
    "tankChar.tanConRooSlo",
    "tankChar.tanDomRooRad",
    "tankChar.conDev",
    "tankChar.conEff",
    "tankChar.tanSha",
    "tankChar.bulTemMet",
    "tankChar.bulTem",
    "tankChar.sheLen2",
    "tankChar.bottomShape",
    "tankChar.bottomSlope",
    "tankChar.liqHeelType",
    "tankChar.liqHeelHeight",
    "tankChar.selSupRoo",
    "tankChar.numCol",
    "tankChar.effColDia",
    "tankChar.intSheCon",
    "tankChar.priSea",
    "tankChar.secSea",
    "tankChar.seaFit",
    "tankChar.decTyp",
    "tankChar.tanCon",
    "tankChar.decCon",
    "tankChar.decSea",
    "tankChar.decConWid",
    "tankChar.decConLen",
    "tankFit.accHatTyp",
    "tankFit.accHatCou",
    "tankFit.colWelTyp",
    "tankFit.colWelCou",
    "tankFit.unsGuiPolTyp",
    "tankFit.unsGuiPolCou",
    "tankFit.sloGuiPolTyp",
    "tankFit.sloGuiPolCou",
    "tankFit.gauFloWelTyp",
    "tankFit.gauFloWelCou",
    "tankFit.gauHatTyp",
    "tankFit.gauHatCou",
    "tankFit.vacBreTyp",
    "tankFit.vacBreCou",
    "tankFit.decDraTyp",
    "tankFit.decDraCou",
    "tankFit.decLegTyp",
    "tankFit.degLegCou",
    "tankFit.fixLegTyp",
    "tankFit.fixLegCou",
    "tankFit.rimVenTyp",
    "tankFit.rimVenCou",
    "tankFit.ladWelTyp",
    "tankFit.ladWelCou",
    "tankFit.ladSloGuiTyp",
    "tankFit.ladSloGuiCon",
    "tankFit.decLegPonTyp",
    "tankFit.decLegPonCou",
    "tankFit.decLegCenTyp",
    "tankFit.decLegCenCou",
    "tankContents.inputType",
    "tankContents.tanCon",
    "tankContents.liqLevMet",
    "tankContents.worLossTurFacMet",
    "tankContents.annData.chemName",
    "tankContents.annData.thr",
    "tankContents.annData.speciation",
    "tankContents.annData.components",
    "tanSolAbs.sheCol",
    "tanSolAbs.sheCon",
    "tanSolAbs.rooCol",
    "tanSolAbs.rooCon",
]

MONTHLY_COLUMNS = ["tankID", "month", "chemName", "thr", "speciation", "components"]
CUSTOM_ORGANIC_COLUMNS = ["Chemical Name", "molWei", "liqDen", "con_A", "con_B", "con_C"]
CUSTOM_MIXTURE_COLUMNS = ["Mixture Name", "Component Name", "Mol Fraction"]
CUSTOM_PETROLEUM_COLUMNS = [
    "Chemical Name",
    "vapMolWei",
    "liqMolWei",
    "liqDen",
    "vapPreEquCon_A",
    "vapPreEquCon_B",
    "crudeOil",
    "componentType",
]
CUSTOM_PETROLEUM_COMPONENT_COLUMNS = ["Chemical Name", "Component Name", "Mol Fraction"]
CUSTOM_MET_COLUMNS = [
    "Location Name",
    "Month",
    "Average Maximum Ambient Temperature (F)",
    "Average Minimum Ambient Temperature (F)",
    "Average Wind Speed (mph)",
    "Average Daily Total Insolation Factor (Btu/ft2/day)",
    "Average Atmospheric Pressure (psi)",
]

INITIAL_TANK_TYPE = {"tanTyp": ""}
INITIAL_TANK_IDENTIFICATION = {
    "tankID": "",
    "tankDescription": "",
    "tankCity": "",
    "tankState": "",
    "company": "",
}
INITIAL_TANK_CHAR = {
    "sheLen": "",
    "sheHei": "",
    "sheDia": "",
    "maxLiqHei": "",
    "avgLiqHei": "",
    "minLiqHei": "",
    "tanHea": "",
    "maxHeaTem": "",
    "avgHeaTem": "",
    "minHeaTem": "",
    "heaCyc": "",
    "rooTyp": "",
    "vacSet": "",
    "preSet": "",
    "vapSpaPre": "",
    "tanIns": "",
    "tanConRooSlo": "",
    "tanDomRooRad": "",
    "conDev": "",
    "conEff": "",
    "tanSha": "",
    "bulTemMet": "",
    "bulTem": "",
    "sheLen2": "",
    "bottomShape": "",
    "bottomSlope": "",
    "liqHeelType": "",
    "liqHeelHeight": "",
    "selSupRoo": "",
    "numCol": "",
    "effColDia": "",
    "intSheCon": "",
    "priSea": "",
    "secSea": "",
    "seaFit": "",
    "decTyp": "",
    "tanCon": "",
    "decCon": "",
    "decSea": "",
    "decConWid": "",
    "decConLen": "",
}
INITIAL_TANK_FIT = {
    "accHatTyp": "",
    "accHatCou": "",
    "colWelTyp": "",
    "colWelCou": "",
    "unsGuiPolTyp": "",
    "unsGuiPolCou": "",
    "sloGuiPolTyp": "",
    "sloGuiPolCou": "",
    "gauFloWelTyp": "",
    "gauFloWelCou": "",
    "gauHatTyp": "",
    "gauHatCou": "",
    "vacBreTyp": "",
    "vacBreCou": "",
    "decDraTyp": "",
    "decDraCou": "",
    "decLegTyp": "",
    "degLegCou": "",
    "fixLegTyp": "",
    "fixLegCou": "",
    "rimVenTyp": "",
    "rimVenCou": "",
    "ladWelTyp": "",
    "ladWelCou": "",
    "ladSloGuiTyp": "",
    "ladSloGuiCon": "",
    "decLegPonTyp": "",
    "decLegPonCou": "",
    "decLegCenTyp": "",
    "decLegCenCou": "",
}
INITIAL_LOCATION = {
    "loc": "",
    "houAvgMinAmbTem": "",
    "houAvgMaxAmbTem": "",
    "avgWinSpe": "",
    "avgDaiTotInsFac": "",
    "avgAtmPre": "",
}
INITIAL_TAN_SOL_ABS = {
    "sheCol": "",
    "sheCon": "",
    "tanSheSurSolAbs": "",
    "rooCol": "",
    "rooCon": "",
    "tanRooSurSolAbs": "",
}
INITIAL_TANK_CONTENTS = {
    "inputType": "",
    "tanCon": "",
    "liqLevMet": "",
    "worLossTurFacMet": "",
    "annData": {"chemName": "", "thr": "", "speciation": "", "components": []},
    "monData": [
        {"month": "January", "chemName": "", "thr": "", "speciation": "", "components": []},
        {"month": "February", "chemName": "", "thr": "", "speciation": "", "components": []},
        {"month": "March", "chemName": "", "thr": "", "speciation": "", "components": []},
        {"month": "April", "chemName": "", "thr": "", "speciation": "", "components": []},
        {"month": "May", "chemName": "", "thr": "", "speciation": "", "components": []},
        {"month": "June", "chemName": "", "thr": "", "speciation": "", "components": []},
        {"month": "July", "chemName": "", "thr": "", "speciation": "", "components": []},
        {"month": "August", "chemName": "", "thr": "", "speciation": "", "components": []},
        {"month": "September", "chemName": "", "thr": "", "speciation": "", "components": []},
        {"month": "October", "chemName": "", "thr": "", "speciation": "", "components": []},
        {"month": "November", "chemName": "", "thr": "", "speciation": "", "components": []},
        {"month": "December", "chemName": "", "thr": "", "speciation": "", "components": []},
    ],
}
INITIAL_PET_CHEM = {"annData": {}, "monData": {}}
INITIAL_PET_DIST = {"annData": {}, "monData": {}}


def deep_merge(base, override):
    if not isinstance(base, dict) or not isinstance(override, dict):
        return copy.deepcopy(override)
    merged = copy.deepcopy(base)
    for key, value in override.items():
        if key in merged and isinstance(merged[key], dict) and isinstance(value, dict):
            merged[key] = deep_merge(merged[key], value)
        else:
            merged[key] = copy.deepcopy(value)
    return merged


def blank_tank(tank_type=""):
    tank = {
        "tankType": copy.deepcopy(INITIAL_TANK_TYPE),
        "tankIdentification": copy.deepcopy(INITIAL_TANK_IDENTIFICATION),
        "location": copy.deepcopy(INITIAL_LOCATION),
        "tankChar": copy.deepcopy(INITIAL_TANK_CHAR),
        "tankFit": copy.deepcopy(INITIAL_TANK_FIT),
        "tankContents": copy.deepcopy(INITIAL_TANK_CONTENTS),
        "tanSolAbs": copy.deepcopy(INITIAL_TAN_SOL_ABS),
        "petChem": copy.deepcopy(INITIAL_PET_CHEM),
        "petDist": copy.deepcopy(INITIAL_PET_DIST),
    }
    if tank_type:
        tank["tankType"]["tanTyp"] = tank_type
    return tank


def default_payload(tank_type="Vertical Fixed Roof Tank"):
    return {
        "tanks": [blank_tank(tank_type=tank_type)],
        "customOrganicLiquids": {},
        "customPetroleumLiquids": {},
        "customMixedOrganicLiquids": {},
        "customLocations": {},
    }


def normalize_payload(payload):
    normalized = {
        "tanks": [],
        "customOrganicLiquids": copy.deepcopy(payload.get("customOrganicLiquids", {})),
        "customPetroleumLiquids": copy.deepcopy(payload.get("customPetroleumLiquids", {})),
        "customMixedOrganicLiquids": copy.deepcopy(payload.get("customMixedOrganicLiquids", {})),
        "customLocations": copy.deepcopy(payload.get("customLocations", {})),
    }
    for raw_tank in payload.get("tanks", []):
        tank = blank_tank()
        for key in TANK_ROW_KEYS:
            if key in raw_tank:
                tank[key] = deep_merge(tank[key], raw_tank[key])
        normalized["tanks"].append(tank)
    return normalized


def dumps_cell(value):
    return json.dumps(value, ensure_ascii=True, separators=(",", ":"))


def append_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def append_json_rows(ws, rows):
    if not rows:
        return
    headers = list(rows[0].keys())
    append_rows(ws, headers, rows)


def build_import_workbook(payload):
    wb = Workbook()
    ws = wb.active
    ws.title = "TankData"
    tank_rows = [{key: dumps_cell(tank[key]) for key in TANK_ROW_KEYS} for tank in payload["tanks"]]
    append_json_rows(ws, tank_rows)
    custom_organic = wb.create_sheet("CustomOrganicLiquids")
    append_rows(
        custom_organic,
        ["Chemical Name", "Json"],
        [
            {"Chemical Name": name, "Json": dumps_cell(value)}
            for name, value in payload["customOrganicLiquids"].items()
        ],
    )
    custom_petroleum = wb.create_sheet("CustomPetroleumLiquids")
    append_rows(
        custom_petroleum,
        ["Chemical Name", "Json"],
        [
            {"Chemical Name": name, "Json": dumps_cell(value)}
            for name, value in payload["customPetroleumLiquids"].items()
        ],
    )
    custom_mixtures = wb.create_sheet("CustomMixtures")
    append_rows(
        custom_mixtures,
        ["Mixture Name", "Json"],
        [
            {"Mixture Name": name, "Json": dumps_cell(value)}
            for name, value in payload["customMixedOrganicLiquids"].items()
        ],
    )
    custom_locations = wb.create_sheet("CustomMeteorologicalData")
    append_rows(
        custom_locations,
        ["Location Name", "Json"],
        [
            {"Location Name": name, "Json": dumps_cell(value)}
            for name, value in payload["customLocations"].items()
        ],
    )
    return wb


def set_nested_value(obj, dotted_key, value):
    parts = dotted_key.split(".")
    cur = obj
    for part in parts[:-1]:
        cur = cur[part]
    cur[parts[-1]] = value


def flatten_fillable_row(tank):
    row = {}
    for key in FILLABLE_TANK_COLUMNS:
        value = tank
        for part in key.split("."):
            value = value[part]
        row[key] = ", ".join(str(item) for item in value) if isinstance(value, list) else value
    return row


def build_fillable_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    lines = [
        "TANKS 5.2 fillable template",
        "",
        "1. Fill the Tanks sheet with one row per tank.",
        "2. Use MonthlyContents only for tanks with tankContents.inputType = Enter Monthly Values.",
        "3. Fill CustomMeteorologicalData with 13 rows per custom location (January-December plus Annual).",
        "4. Convert this workbook with:",
        "   python tools\\tanks52_workbook.py --from-fillable INPUT.xlsx OUTPUT_import.xlsx",
    ]
    for idx, line in enumerate(lines, start=1):
        ws.cell(row=idx, column=1, value=line)
    ws.column_dimensions["A"].width = 120

    tanks = wb.create_sheet("Tanks")
    append_rows(tanks, FILLABLE_TANK_COLUMNS, [flatten_fillable_row(blank_tank("Vertical Fixed Roof Tank"))])
    tanks.freeze_panes = "A2"

    monthly = wb.create_sheet("MonthlyContents")
    append_rows(monthly, MONTHLY_COLUMNS, [])
    monthly.freeze_panes = "A2"

    custom_met = wb.create_sheet("CustomMeteorologicalData")
    append_rows(custom_met, CUSTOM_MET_COLUMNS, [])
    custom_met.freeze_panes = "A2"

    custom_org = wb.create_sheet("CustomOrganicLiquids")
    append_rows(custom_org, CUSTOM_ORGANIC_COLUMNS, [])
    custom_org.freeze_panes = "A2"

    custom_mix = wb.create_sheet("CustomMixtures")
    append_rows(custom_mix, CUSTOM_MIXTURE_COLUMNS, [])
    custom_mix.freeze_panes = "A2"

    custom_pet = wb.create_sheet("CustomPetroleumLiquids")
    append_rows(custom_pet, CUSTOM_PETROLEUM_COLUMNS, [])
    custom_pet.freeze_panes = "A2"

    custom_pet_comp = wb.create_sheet("CustomPetroleumComponents")
    append_rows(custom_pet_comp, CUSTOM_PETROLEUM_COMPONENT_COLUMNS, [])
    custom_pet_comp.freeze_panes = "A2"

    reference = wb.create_sheet("Reference")
    reference.append(["Useful values"])
    reference["A1"].font = Font(bold=True)
    reference.append(["Months"])
    for month in MONTH_LABELS:
        reference.append([month])
    reference.append([])
    reference.append(["Common tank types"])
    for value in [
        "Vertical Fixed Roof Tank",
        "Horizontal Fixed Roof Tank",
        "Internal Floating Roof Tank",
        "External Floating Roof Tank",
        "Domed External Floating Roof Tank",
    ]:
        reference.append([value])
    reference.column_dimensions["A"].width = 40

    for ws_item in wb.worksheets:
        if ws_item.title == "Instructions":
            continue
        for cell in ws_item[1]:
            cell.font = Font(bold=True)
        ws_item.auto_filter.ref = ws_item.dimensions
    return wb


def read_sheet_dicts(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        return []
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    data = []
    for raw in rows[1:]:
        if not any(v not in (None, "") for v in raw):
            continue
        row = {}
        for idx, header in enumerate(headers):
            if header:
                row[header] = raw[idx] if idx < len(raw) else None
        data.append(row)
    return data


def coerce_scalar(value):
    if value is None:
        return ""
    return value.strip() if isinstance(value, str) else value


def split_components(value):
    text = coerce_scalar(value)
    if not text:
        return []
    return [item.strip() for item in str(text).split(",") if item.strip()]


def maybe_number(value):
    value = coerce_scalar(value)
    if value == "":
        return ""
    if isinstance(value, (int, float)):
        return value
    try:
        number = float(value)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def maybe_bool(value):
    value = coerce_scalar(value)
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "yes", "y", "1"}:
        return True
    if text in {"false", "no", "n", "0", ""}:
        return False
    return value


@lru_cache(maxsize=1)
def load_source_map():
    if not SOURCE_MAP_PATH.exists():
        raise FileNotFoundError(f"Missing required EPA source map: {SOURCE_MAP_PATH}")
    return json.loads(SOURCE_MAP_PATH.read_text(encoding="utf-8", errors="ignore"))


def js_source_to_json(source_text):
    text = source_text.replace("export default", "", 1).strip()
    if text.endswith(";"):
        text = text[:-1]
    text = re.sub(r"/\*.*?\*/", "", text, flags=re.S)
    text = re.sub(r"//.*?$", "", text, flags=re.M)
    text = re.sub(r"\bNaN\b", "null", text)
    text = re.sub(r"'([^'\\]*(?:\\.[^'\\]*)*)'", lambda m: json.dumps(m.group(1)), text)
    text = re.sub(r",\s*([}\]])", r"\1", text)
    return json.loads(text)


@lru_cache(maxsize=None)
def load_app_table(source_name):
    source_map = load_source_map()
    idx = source_map["sources"].index(source_name)
    return js_source_to_json(source_map["sourcesContent"][idx])


def ap42_locations():
    return load_app_table("tables/Table7_1_7.js")["Table7_1_7"]["metDat"]


def solar_absorptance_table():
    return load_app_table("tables/Table7_1_6.js")["Table7_1_6"]["paiSolAbs"]


def builtin_petroleum_liquids():
    return load_app_table("tables/Table7_1_2.js")["Table7_1_2"]["petroleumLiquids"]


def builtin_organic_liquids():
    return load_app_table("tables/Table7_1_3.js")["Table7_1_3"]["petrochemicals"]


def lookup_solar_abs(color, condition):
    if not color or not condition:
        return ""
    entry = solar_absorptance_table().get(color)
    if not entry:
        return ""
    if condition == "New":
        return entry["refConNew"]
    if condition == "Average":
        return entry["refConAvg"]
    return entry["refConAge"]


def monthly_array_from_location_dict(location_dict, key):
    values = location_dict[key]
    if isinstance(values, dict):
        return [values[label] for label in MONTH_SHORT]
    return list(values)


def build_custom_locations(sheet_rows):
    grouped = defaultdict(dict)
    for row in sheet_rows:
        name = coerce_scalar(row.get("Location Name"))
        month = coerce_scalar(row.get("Month"))
        if not name or not month:
            continue
        month_idx = MONTH_TO_INDEX.get(month.lower())
        if month_idx is None:
            raise ValueError(f"Unknown custom meteorological month '{month}' for '{name}'.")
        grouped[name][month_idx] = {
            "maxTemp": maybe_number(row.get("Average Maximum Ambient Temperature (F)")),
            "minTemp": maybe_number(row.get("Average Minimum Ambient Temperature (F)")),
            "windSpeed": maybe_number(row.get("Average Wind Speed (mph)")),
            "avgDaiTotInsFac": maybe_number(row.get("Average Daily Total Insolation Factor (Btu/ft2/day)")),
            "pressure": maybe_number(row.get("Average Atmospheric Pressure (psi)")),
        }
    custom_locations = {}
    for name, mapping in grouped.items():
        missing = [MONTH_LABELS[idx] for idx in range(13) if idx not in mapping]
        if missing:
            raise ValueError(f"Custom meteorological location '{name}' is missing rows for: {', '.join(missing)}.")
        arr = []
        for idx in range(13):
            item = mapping[idx]
            arr.append(
                {
                    "maxTemp": item["maxTemp"],
                    "minTemp": item["minTemp"],
                    "windSpeed": item["windSpeed"],
                    "avgDaiTotInsFac": item["avgDaiTotInsFac"],
                    "pressure": item["pressure"] if idx == 12 else "",
                }
            )
        custom_locations[name] = arr
    return custom_locations


def build_custom_organic_liquids(rows):
    result = {}
    for row in rows:
        name = coerce_scalar(row.get("Chemical Name"))
        if name:
            result[name] = {
                "molWei": maybe_number(row.get("molWei")),
                "liqDen": maybe_number(row.get("liqDen")),
                "con_A": maybe_number(row.get("con_A")),
                "con_B": maybe_number(row.get("con_B")),
                "con_C": maybe_number(row.get("con_C")),
            }
    return result


def build_custom_mixtures(rows, organic_lookup):
    grouped = defaultdict(list)
    for row in rows:
        mixture_name = coerce_scalar(row.get("Mixture Name"))
        component_name = coerce_scalar(row.get("Component Name"))
        mol_frac = maybe_number(row.get("Mol Fraction"))
        if not mixture_name or not component_name:
            continue
        if component_name not in organic_lookup:
            raise ValueError(f"Custom mixture '{mixture_name}' references unknown component '{component_name}'.")
        component = organic_lookup[component_name]
        grouped[mixture_name].append(
            {
                "chemName": component_name,
                "molFrac": mol_frac,
                "molWei": maybe_number(component.get("molWei")),
                "liqDen": maybe_number(component.get("liqDen")),
                "con_A": maybe_number(component.get("con_A")),
                "con_B": maybe_number(component.get("con_B")),
                "con_C": maybe_number(component.get("con_C")),
            }
        )
    return dict(grouped)


def build_custom_petroleum_liquids(rows, component_rows, organic_lookup):
    grouped_components = defaultdict(list)
    for row in component_rows:
        liquid_name = coerce_scalar(row.get("Chemical Name"))
        component_name = coerce_scalar(row.get("Component Name"))
        mol_frac = maybe_number(row.get("Mol Fraction"))
        if not liquid_name or not component_name:
            continue
        if component_name not in organic_lookup:
            raise ValueError(f"Custom petroleum liquid '{liquid_name}' references unknown component '{component_name}'.")
        grouped_components[liquid_name].append({"chemName": component_name, "molFrac": mol_frac})

    result = {}
    for row in rows:
        name = coerce_scalar(row.get("Chemical Name"))
        if not name:
            continue
        entry = {
            "vapMolWei": maybe_number(row.get("vapMolWei")),
            "liqMolWei": maybe_number(row.get("liqMolWei")),
            "liqDen": maybe_number(row.get("liqDen")),
            "vapPreEquCon_A": maybe_number(row.get("vapPreEquCon_A")),
            "vapPreEquCon_B": maybe_number(row.get("vapPreEquCon_B")),
            "crudeOil": maybe_bool(row.get("crudeOil")),
        }
        component_type = coerce_scalar(row.get("componentType"))
        if grouped_components.get(name):
            entry["componentData"] = {
                "type": component_type or "Full Speciation",
                "components": grouped_components[name],
            }
        result[name] = entry
    return result


def build_location(loc_name, custom_locations):
    raw_name = coerce_scalar(loc_name)
    if not raw_name:
        return copy.deepcopy(INITIAL_LOCATION)

    name = raw_name[8:] if raw_name.startswith("Custom: ") else raw_name
    if name in custom_locations:
        arr = custom_locations[name]
        return {
            "loc": f"Custom: {name}",
            "houAvgMinAmbTem": [entry["minTemp"] for entry in arr],
            "houAvgMaxAmbTem": [entry["maxTemp"] for entry in arr],
            "avgWinSpe": [entry["windSpeed"] for entry in arr],
            "avgDaiTotInsFac": [entry["avgDaiTotInsFac"] for entry in arr],
            "avgAtmPre": arr[12]["pressure"],
        }

    built_in = ap42_locations().get(raw_name)
    if built_in is None:
        raise ValueError(f"Unknown meteorological location '{raw_name}'.")
    return {
        "loc": raw_name,
        "houAvgMinAmbTem": monthly_array_from_location_dict(built_in, "houAvgMinAmbTem"),
        "houAvgMaxAmbTem": monthly_array_from_location_dict(built_in, "houAvgMaxAmbTem"),
        "avgWinSpe": monthly_array_from_location_dict(built_in, "avgWinSpe"),
        "avgDaiTotInsFac": monthly_array_from_location_dict(built_in, "avgDaiTotInsFac"),
        "avgAtmPre": built_in["avgAtmPre"],
    }


def normalize_organic_entry(name, lookup):
    source = lookup[name]
    return {
        "chemName": name,
        "molWei": maybe_number(source.get("molWei")),
        "liqDen": maybe_number(source.get("liqDen")),
        "con_A": maybe_number(source.get("con_A")),
        "con_B": maybe_number(source.get("con_B")),
        "con_C": maybe_number(source.get("con_C")),
    }


def normalize_petroleum_entry(name, lookup):
    source = lookup[name]
    entry = {
        "chemName": name,
        "vapMolWei": maybe_number(source.get("vapMolWei")),
        "liqMolWei": maybe_number(source.get("liqMolWei")),
        "liqDen": maybe_number(source.get("liqDen")),
        "vapPreEquCon_A": maybe_number(source.get("vapPreEquCon_A")),
        "vapPreEquCon_B": maybe_number(source.get("vapPreEquCon_B")),
    }
    if "crudeOil" in source:
        entry["crudeOil"] = source.get("crudeOil")
    return entry


def build_annual_content_data(tank, organic_lookup, mixture_lookup, petroleum_lookup, custom_petroleum_lookup):
    contents_type = tank["tankContents"]["tanCon"]
    chem_name = tank["tankContents"]["annData"]["chemName"]
    if not chem_name:
        return

    if contents_type in {"Organic Liquids", "Custom Organic Liquids"}:
        tank["petChem"]["annData"] = normalize_organic_entry(chem_name, organic_lookup)
        tank["petChem"]["monData"] = {}
        tank["petDist"]["annData"] = {}
        tank["petDist"]["monData"] = {}
        return

    if contents_type == "Custom Organic Mixtures":
        mixture = mixture_lookup[chem_name]
        annual_mix = {}
        for idx, component in enumerate(mixture):
            annual_mix[idx] = {
                "chemName": component["chemName"],
                "molWei": maybe_number(component["molWei"]),
                "molFrac": maybe_number(component["molFrac"]),
                "liqDen": maybe_number(component["liqDen"]),
                "con_A": maybe_number(component["con_A"]),
                "con_B": maybe_number(component["con_B"]),
                "con_C": maybe_number(component["con_C"]),
            }
        annual_mix["length"] = len(mixture)
        tank["petChem"]["annData"] = annual_mix
        tank["petChem"]["monData"] = {}
        tank["petDist"]["annData"] = {}
        tank["petDist"]["monData"] = {}
        return

    if contents_type in {"Petroleum Liquids", "Custom Petroleum Liquids"}:
        petroleum_source = custom_petroleum_lookup if contents_type == "Custom Petroleum Liquids" else petroleum_lookup
        tank["petDist"]["annData"] = normalize_petroleum_entry(chem_name, petroleum_source)
        tank["petDist"]["monData"] = {}
        tank["petChem"]["monData"] = {}
        tank["petChem"]["annData"] = {}
        if contents_type == "Custom Petroleum Liquids":
            component_data = petroleum_source[chem_name].get("componentData")
            if component_data:
                pet_chem_components = {}
                for idx, component in enumerate(component_data["components"]):
                    source = organic_lookup[component["chemName"]]
                    pet_chem_components[idx] = {
                        "chemName": component["chemName"],
                        "molFrac": maybe_number(component["molFrac"]),
                        "molWei": maybe_number(source.get("molWei")),
                        "liqDen": maybe_number(source.get("liqDen")),
                        "con_A": maybe_number(source.get("con_A")),
                        "con_B": maybe_number(source.get("con_B")),
                        "con_C": maybe_number(source.get("con_C")),
                    }
                pet_chem_components["type"] = component_data.get("type", "Full Speciation")
                pet_chem_components["length"] = len(component_data["components"])
                tank["petChem"]["annData"] = pet_chem_components
        return

    raise ValueError(f"Unsupported tank contents type '{contents_type}'.")


def build_monthly_content_data(tank, monthly_rows, organic_lookup, mixture_lookup, petroleum_lookup, custom_petroleum_lookup):
    contents_type = tank["tankContents"]["tanCon"]
    mon_data = copy.deepcopy(INITIAL_TANK_CONTENTS["monData"])
    rows_by_idx = {}
    for row in monthly_rows:
        month = coerce_scalar(row.get("month"))
        idx = MONTH_TO_INDEX.get(month.lower()) if month else None
        if idx is None or idx > 11:
            raise ValueError(f"Invalid monthly contents month '{month}' for tank '{tank['tankIdentification']['tankID']}'.")
        rows_by_idx[idx] = row

    for idx, item in enumerate(mon_data):
        row = rows_by_idx.get(idx, {})
        item["chemName"] = coerce_scalar(row.get("chemName"))
        item["thr"] = maybe_number(row.get("thr"))
        item["speciation"] = coerce_scalar(row.get("speciation"))
        item["components"] = split_components(row.get("components"))
    tank["tankContents"]["monData"] = mon_data

    if contents_type in {"Organic Liquids", "Custom Organic Liquids"}:
        monthly = {"chemName": [], "molWei": [], "liqDen": [], "con_A": [], "con_B": [], "con_C": []}
        for item in mon_data:
            if not item["chemName"]:
                for key in monthly:
                    monthly[key].append("")
                continue
            source = organic_lookup[item["chemName"]]
            monthly["chemName"].append(item["chemName"])
            monthly["molWei"].append(maybe_number(source.get("molWei")))
            monthly["liqDen"].append(maybe_number(source.get("liqDen")))
            monthly["con_A"].append(maybe_number(source.get("con_A")))
            monthly["con_B"].append(maybe_number(source.get("con_B")))
            monthly["con_C"].append(maybe_number(source.get("con_C")))
        tank["petChem"]["monData"] = monthly
        tank["petChem"]["annData"] = {}
        tank["petDist"]["annData"] = {}
        tank["petDist"]["monData"] = {}
        return

    if contents_type == "Custom Organic Mixtures":
        monthly_mix = {}
        for idx, item in enumerate(mon_data):
            if not item["chemName"]:
                monthly_mix[idx] = {"chemName": "", "molWei": "", "molFrac": "", "liqDen": "", "con_A": "", "con_B": "", "con_C": ""}
                continue
            mixture = mixture_lookup[item["chemName"]]
            month_entry = {}
            for comp_idx, component in enumerate(mixture):
                month_entry[comp_idx] = {
                    "chemName": component["chemName"],
                    "molWei": maybe_number(component["molWei"]),
                    "molFrac": maybe_number(component["molFrac"]),
                    "liqDen": maybe_number(component["liqDen"]),
                    "con_A": maybe_number(component["con_A"]),
                    "con_B": maybe_number(component["con_B"]),
                    "con_C": maybe_number(component["con_C"]),
                }
            month_entry["length"] = len(mixture)
            monthly_mix[idx] = month_entry
        tank["petChem"]["monData"] = monthly_mix
        tank["petChem"]["annData"] = {}
        tank["petDist"]["annData"] = {}
        tank["petDist"]["monData"] = {}
        return

    if contents_type in {"Petroleum Liquids", "Custom Petroleum Liquids"}:
        petroleum_source = custom_petroleum_lookup if contents_type == "Custom Petroleum Liquids" else petroleum_lookup
        monthly = {
            "chemName": [],
            "vapMolWei": [],
            "liqDen": [],
            "liqMolWei": [],
            "vapPreEquCon_A": [],
            "vapPreEquCon_B": [],
            "crudeOil": [],
        }
        pet_chem_monthly = {}
        for idx, item in enumerate(mon_data):
            if not item["chemName"]:
                for key in monthly:
                    monthly[key].append("")
                if contents_type == "Custom Petroleum Liquids":
                    pet_chem_monthly[idx] = {}
                continue
            source = petroleum_source[item["chemName"]]
            monthly["chemName"].append(item["chemName"])
            monthly["vapMolWei"].append(maybe_number(source.get("vapMolWei")))
            monthly["liqDen"].append(maybe_number(source.get("liqDen")))
            monthly["liqMolWei"].append(maybe_number(source.get("liqMolWei")))
            monthly["vapPreEquCon_A"].append(maybe_number(source.get("vapPreEquCon_A")))
            monthly["vapPreEquCon_B"].append(maybe_number(source.get("vapPreEquCon_B")))
            monthly["crudeOil"].append(source.get("crudeOil", False))
            if contents_type == "Custom Petroleum Liquids":
                component_data = source.get("componentData")
                if component_data:
                    month_entry = {}
                    for comp_idx, component in enumerate(component_data["components"]):
                        organic_source = organic_lookup[component["chemName"]]
                        month_entry[comp_idx] = {
                            "chemName": component["chemName"],
                            "molFrac": maybe_number(component["molFrac"]),
                            "molWei": maybe_number(organic_source.get("molWei")),
                            "liqDen": maybe_number(organic_source.get("liqDen")),
                            "con_A": maybe_number(organic_source.get("con_A")),
                            "con_B": maybe_number(organic_source.get("con_B")),
                            "con_C": maybe_number(organic_source.get("con_C")),
                        }
                    month_entry["type"] = component_data.get("type", "Full Speciation")
                    month_entry["length"] = len(component_data["components"])
                    pet_chem_monthly[idx] = month_entry
                else:
                    pet_chem_monthly[idx] = {}
        tank["petDist"]["monData"] = monthly
        tank["petDist"]["annData"] = {}
        tank["petChem"]["annData"] = {}
        tank["petChem"]["monData"] = pet_chem_monthly if contents_type == "Custom Petroleum Liquids" else {}
        return

    raise ValueError(f"Unsupported tank contents type '{contents_type}'.")


def payload_from_fillable_workbook(path):
    wb = load_workbook(path)
    tank_rows = read_sheet_dicts(wb, "Tanks")
    monthly_rows = read_sheet_dicts(wb, "MonthlyContents")
    custom_met_rows = read_sheet_dicts(wb, "CustomMeteorologicalData")
    custom_org_rows = read_sheet_dicts(wb, "CustomOrganicLiquids")
    custom_mix_rows = read_sheet_dicts(wb, "CustomMixtures")
    custom_pet_rows = read_sheet_dicts(wb, "CustomPetroleumLiquids")
    custom_pet_comp_rows = read_sheet_dicts(wb, "CustomPetroleumComponents")

    custom_locations = build_custom_locations(custom_met_rows)
    custom_organic = build_custom_organic_liquids(custom_org_rows)
    organic_lookup = {}
    organic_lookup.update(builtin_organic_liquids())
    organic_lookup.update(custom_organic)
    custom_mixtures = build_custom_mixtures(custom_mix_rows, organic_lookup)
    custom_petroleum = build_custom_petroleum_liquids(custom_pet_rows, custom_pet_comp_rows, organic_lookup)
    petroleum_lookup = builtin_petroleum_liquids()

    monthly_by_tank = defaultdict(list)
    for row in monthly_rows:
        tank_id = coerce_scalar(row.get("tankID"))
        if tank_id:
            monthly_by_tank[tank_id].append(row)

    tanks = []
    for row in tank_rows:
        tank = blank_tank()
        for key in FILLABLE_TANK_COLUMNS:
            value = row.get(key)
            value = split_components(value) if key.endswith(".components") else maybe_number(value)
            set_nested_value(tank, key, value)

        tank_id = tank["tankIdentification"]["tankID"]
        tank["location"] = build_location(tank["location"]["loc"], custom_locations)
        tank["tanSolAbs"]["tanSheSurSolAbs"] = lookup_solar_abs(tank["tanSolAbs"]["sheCol"], tank["tanSolAbs"]["sheCon"])
        tank["tanSolAbs"]["tanRooSurSolAbs"] = lookup_solar_abs(tank["tanSolAbs"]["rooCol"], tank["tanSolAbs"]["rooCon"])

        if tank["tankContents"]["inputType"] == "Enter Monthly Values":
            build_monthly_content_data(tank, monthly_by_tank.get(tank_id, []), organic_lookup, custom_mixtures, petroleum_lookup, custom_petroleum)
        else:
            build_annual_content_data(tank, organic_lookup, custom_mixtures, petroleum_lookup, custom_petroleum)
        tanks.append(tank)

    return normalize_payload(
        {
            "tanks": tanks,
            "customOrganicLiquids": custom_organic,
            "customPetroleumLiquids": custom_petroleum,
            "customMixedOrganicLiquids": custom_mixtures,
            "customLocations": custom_locations,
        }
    )


def parse_args():
    parser = argparse.ArgumentParser(description="Build or convert TANKS 5.2 workbooks.")
    parser.add_argument("input_json", nargs="?", help="Path to input JSON payload.")
    parser.add_argument("output_xlsx", nargs="?", help="Path to output XLSX workbook.")
    parser.add_argument("--template-out", help="Write a starter JSON payload.")
    parser.add_argument("--fillable-template-out", help="Write a human-editable XLSX template.")
    parser.add_argument(
        "--from-fillable",
        nargs=2,
        metavar=("INPUT_FILLABLE_XLSX", "OUTPUT_IMPORT_XLSX"),
        help="Convert a fillable workbook into a TANKS import workbook.",
    )
    parser.add_argument(
        "--tank-type",
        default="Vertical Fixed Roof Tank",
        help="Tank type to use in the starter JSON template.",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    if args.template_out:
        Path(args.template_out).write_text(json.dumps(default_payload(tank_type=args.tank_type), indent=2), encoding="utf-8")
        return

    if args.fillable_template_out:
        build_fillable_template_workbook().save(args.fillable_template_out)
        return

    if args.from_fillable:
        input_path, output_path = args.from_fillable
        payload = payload_from_fillable_workbook(input_path)
        build_import_workbook(payload).save(output_path)
        return

    if not args.input_json or not args.output_xlsx:
        raise SystemExit("Provide INPUT_JSON and OUTPUT_XLSX, or use one of the template/conversion flags.")

    payload = json.loads(Path(args.input_json).read_text(encoding="utf-8"))
    build_import_workbook(normalize_payload(payload)).save(args.output_xlsx)


if __name__ == "__main__":
    main()
